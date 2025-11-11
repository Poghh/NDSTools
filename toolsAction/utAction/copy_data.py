import os
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font
from copy import copy


def copy_data_action(test_file_path, doc_file_path, progress_callback=None):
    """
    Copy data from 項目一覧 sheet in mocks file to 項目一覧 sheet in unit test file
    
    Args:
        test_file_path (str): Path to the mocks file (source)
        doc_file_path (str): Path to the unit test file (destination)
        progress_callback (function): Optional callback to report progress
        
    Returns:
        dict: Result information
    """
    try:
        if progress_callback:
            progress_callback("Đang kiểm tra file...")
        
        # Validate files exist
        if not os.path.exists(test_file_path):
            raise FileNotFoundError(f"Không tìm thấy file mocks: {test_file_path}")
        
        if not os.path.exists(doc_file_path):
            raise FileNotFoundError(f"Không tìm thấy file unit test: {doc_file_path}")
        
        # Check if both files are Excel files
        test_ext = os.path.splitext(test_file_path)[1].lower()
        doc_ext = os.path.splitext(doc_file_path)[1].lower()
        
        if test_ext not in ['.xlsx', '.xls'] or doc_ext not in ['.xlsx', '.xls']:
            raise ValueError("Cả hai file phải là file Excel (.xlsx hoặc .xls)")
        
        if progress_callback:
            progress_callback("Đang đọc file mocks...")
        
        # Load the mocks file (source)
        source_wb = load_workbook(test_file_path, data_only=True)
        
        # Find the 項目一覧 sheet in mocks file
        sheet_name = "項目一覧"
        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"Không tìm thấy sheet '{sheet_name}' trong file mocks")
        
        source_ws = source_wb[sheet_name]
        
        if progress_callback:
            progress_callback("Đang đọc dữ liệu từ sheet 項目一覧...")
        
        # Get all cells with data and formatting from source sheet (starting from row 5)
        max_row = source_ws.max_row
        max_col = source_ws.max_column
        start_source_row = 5  # Copy from row 5 onwards
        
        # Store source cells for copying with formatting (from row 5)
        source_cells = []
        for row in range(start_source_row, max_row + 1):
            row_cells = []
            for col in range(1, max_col + 1):
                source_cell = source_ws.cell(row=row, column=col)
                row_cells.append(source_cell)
            source_cells.append(row_cells)
        
        if progress_callback:
            progress_callback("Đang mở file unit test...")
        
        # Load the unit test file (destination)
        dest_wb = load_workbook(doc_file_path)
        
        # Find or create the 項目一覧 sheet in unit test file
        if sheet_name not in dest_wb.sheetnames:
            dest_ws = dest_wb.create_sheet(sheet_name)
        else:
            dest_ws = dest_wb[sheet_name]
        
        if progress_callback:
            progress_callback("Đang sao chép dữ liệu và formatting...")
        
        # Paste data starting from row 6 with full formatting
        start_row = 6
        
        # Copy merged cell ranges information first
        merged_ranges = []
        for merged_range in source_ws.merged_cells.ranges:
            # Calculate the offset for the merged range in destination
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            
            # Only include merged ranges that are within our copying area (from row 5 onwards)
            if min_row >= start_source_row:
                new_min_row = start_row + (min_row - start_source_row)
                new_max_row = start_row + (max_row - start_source_row)
                merged_ranges.append((new_min_row, new_max_row, min_col, max_col))
        
        # Copy cell data and formatting
        for row_idx, row_cells in enumerate(source_cells):
            for col_idx, source_cell in enumerate(row_cells):
                dest_cell = dest_ws.cell(
                    row=start_row + row_idx, 
                    column=col_idx + 1
                )
                
                # Copy cell value
                dest_cell.value = source_cell.value
                
                # Copy cell formatting if source cell has formatting
                if source_cell.has_style:
                    dest_cell.font = copy(source_cell.font)
                    dest_cell.border = copy(source_cell.border)
                    dest_cell.fill = copy(source_cell.fill)
                    dest_cell.number_format = source_cell.number_format
                    dest_cell.protection = copy(source_cell.protection)
                    dest_cell.alignment = copy(source_cell.alignment)
        
        # Apply merged cell ranges to destination
        for new_min_row, new_max_row, min_col, max_col in merged_ranges:
            try:
                dest_ws.merge_cells(
                    start_row=new_min_row,
                    start_column=min_col,
                    end_row=new_max_row,
                    end_column=max_col
                )
            except Exception as e:
                print(f"Warning: Could not merge cells {new_min_row}:{min_col} to {new_max_row}:{max_col} - {e}")

        if progress_callback:
            progress_callback("Đang thêm border và text bổ sung...")
        
        # Add outside border around the copied data
        if source_cells:
            # Define border style
            thin_border = Side(style='thin', color='000000')
            border_style = Border(
                left=thin_border,
                right=thin_border,
                top=thin_border,
                bottom=thin_border
            )
            
            # Calculate the range of copied data
            end_row = start_row + len(source_cells) - 1
            end_col = max_col
            
            # Apply border to the outer cells
            # Top border
            for col in range(1, end_col + 1):
                cell = dest_ws.cell(row=start_row, column=col)
                new_border = Border(
                    top=thin_border,
                    left=cell.border.left,
                    right=cell.border.right,
                    bottom=cell.border.bottom
                )
                if col == 1:  # Left-most cell
                    new_border = Border(
                        top=thin_border,
                        left=thin_border,
                        right=cell.border.right,
                        bottom=cell.border.bottom
                    )
                if col == end_col:  # Right-most cell
                    new_border = Border(
                        top=thin_border,
                        left=cell.border.left,
                        right=thin_border,
                        bottom=cell.border.bottom
                    )
                cell.border = new_border
            
            # Bottom border
            for col in range(1, end_col + 1):
                cell = dest_ws.cell(row=end_row, column=col)
                new_border = Border(
                    bottom=thin_border,
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top
                )
                if col == 1:  # Left-most cell
                    new_border = Border(
                        bottom=thin_border,
                        left=thin_border,
                        right=cell.border.right,
                        top=cell.border.top
                    )
                if col == end_col:  # Right-most cell
                    new_border = Border(
                        bottom=thin_border,
                        left=cell.border.left,
                        right=thin_border,
                        top=cell.border.top
                    )
                cell.border = new_border
            
            # Left border
            for row in range(start_row, end_row + 1):
                cell = dest_ws.cell(row=row, column=1)
                new_border = Border(
                    left=thin_border,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                cell.border = new_border
            
            # Right border
            for row in range(start_row, end_row + 1):
                cell = dest_ws.cell(row=row, column=end_col)
                new_border = Border(
                    right=thin_border,
                    left=cell.border.left,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                cell.border = new_border
            
            # Copy data from "data_1" sheet to the next column after the last copied column, starting from row 5
            data_1_sheet_name = "data_1"
            if data_1_sheet_name in dest_wb.sheetnames:
                if progress_callback:
                    progress_callback("Đang copy dữ liệu từ sheet data_1...")
                
                data_1_ws = dest_wb[data_1_sheet_name]
                data_1_max_row = data_1_ws.max_row
                data_1_max_col = data_1_ws.max_column
                
                # Copy all data from data_1 sheet with formatting
                paste_start_col = end_col + 1
                paste_start_row = 5
                
                # Copy merged cell ranges from data_1 sheet
                data_1_merged_ranges = []
                for merged_range in data_1_ws.merged_cells.ranges:
                    min_row = merged_range.min_row
                    max_row = merged_range.max_row
                    min_col = merged_range.min_col
                    max_col = merged_range.max_col
                    
                    # Calculate new position in destination
                    new_min_row = paste_start_row + min_row - 1
                    new_max_row = paste_start_row + max_row - 1
                    new_min_col = paste_start_col + min_col - 1
                    new_max_col = paste_start_col + max_col - 1
                    data_1_merged_ranges.append((new_min_row, new_max_row, new_min_col, new_max_col))
                
                # Copy cell data and formatting
                for row in range(1, data_1_max_row + 1):
                    for col in range(1, data_1_max_col + 1):
                        source_cell = data_1_ws.cell(row=row, column=col)
                        dest_cell = dest_ws.cell(
                            row=paste_start_row + row - 1,
                            column=paste_start_col + col - 1
                        )
                        
                        # Copy cell value
                        dest_cell.value = source_cell.value
                        
                        # Copy cell formatting if source cell has formatting
                        if source_cell.has_style:
                            dest_cell.font = copy(source_cell.font)
                            dest_cell.border = copy(source_cell.border)
                            dest_cell.fill = copy(source_cell.fill)
                            dest_cell.number_format = source_cell.number_format
                            dest_cell.protection = copy(source_cell.protection)
                            dest_cell.alignment = copy(source_cell.alignment)
                
                # Apply merged cell ranges for data_1
                for new_min_row, new_max_row, new_min_col, new_max_col in data_1_merged_ranges:
                    try:
                        dest_ws.merge_cells(
                            start_row=new_min_row,
                            start_column=new_min_col,
                            end_row=new_max_row,
                            end_column=new_max_col
                        )
                    except Exception as e:
                        print(f"Warning: Could not merge data_1 cells {new_min_row}:{new_min_col} to {new_max_row}:{new_max_col} - {e}")
                
                # Add table borders below data_1 section
                if progress_callback:
                    progress_callback("Đang thêm border cho bảng dưới data_1...")
                
                # Calculate table dimensions
                table_start_row = 9  # Fixed start row
                table_end_row = start_row + len(source_cells) - 1  # Last row of mocks data
                table_start_col = paste_start_col  # First column after mocks data
                table_end_col = paste_start_col + data_1_max_col - 1  # Last column of data_1
                
                # Create thin border for the table area
                table_border = Side(style='thin', color='000000')
                
                # Add borders to all cells in the table area
                for row in range(table_start_row, table_end_row + 1):
                    for col in range(table_start_col, table_end_col + 1):
                        cell = dest_ws.cell(row=row, column=col)
                        
                        # Add thin borders to all sides of every cell to create a grid
                        new_border = Border(
                            top=table_border,
                            bottom=table_border,
                            left=table_border,
                            right=table_border
                        )
                        cell.border = new_border
                
                # Delete data_1 sheet after copying its data
                if progress_callback:
                    progress_callback("Đang xóa sheet data_1...")
                
                try:
                    dest_wb.remove(data_1_ws)
                except Exception as e:
                    print(f"Warning: Could not delete data_1 sheet - {e}")
            else:
                # Fallback: Add "Update sau" text if data_1 sheet doesn't exist
                update_cell = dest_ws.cell(row=5, column=end_col + 1)
                update_cell.value = "Update sau"
        
        if progress_callback:
            progress_callback("Đang lưu file...")
        
        # Save the destination file
        dest_wb.save(doc_file_path)
        
        # Close workbooks
        source_wb.close()
        dest_wb.close()
        
        result = {
            "test_file": os.path.basename(test_file_path),
            "doc_file": os.path.basename(doc_file_path),
            "sheet_name": sheet_name,
            "rows_copied": len(source_cells),
            "cols_copied": max_col if source_cells else 0,
            "source_start_row": start_source_row,
            "paste_start_row": start_row,
            "has_border": True,
            "data_1_copied": data_1_sheet_name in dest_wb.sheetnames if source_cells else False,
            "data_1_start_col": max_col + 1 if source_cells else 0,
            "data_1_start_row": 5,
            "data_1_sheet_deleted": True,
            "table_borders_added": True,
            "table_start_row": 9,
            "table_end_row": start_row + len(source_cells) - 1 if source_cells else 0,
            "status": "success"
        }
        
        if progress_callback:
            progress_callback("Hoàn thành!")
        
        return result
        
    except Exception as e:
        if progress_callback:
            progress_callback(f"Lỗi: {str(e)}")
        raise e


def copy_action_data(test_file_path, doc_file_path, progress_callback=None):
    """
    Copy data from アクション一覧 sheet in mocks file to アクション一覧 sheet in unit test file
    Find "アクションNo." column and copy all data below it
    
    Args:
        test_file_path (str): Path to the mocks file (source)
        doc_file_path (str): Path to the unit test file (destination)
        progress_callback (function): Optional callback to report progress
        
    Returns:
        dict: Result information
    """
    try:
        if progress_callback:
            progress_callback("Đang kiểm tra file...")
        
        # Validate files exist
        if not os.path.exists(test_file_path):
            raise FileNotFoundError(f"Không tìm thấy file mocks: {test_file_path}")
        
        if not os.path.exists(doc_file_path):
            raise FileNotFoundError(f"Không tìm thấy file unit test: {doc_file_path}")
        
        # Check if both files are Excel files
        test_ext = os.path.splitext(test_file_path)[1].lower()
        doc_ext = os.path.splitext(doc_file_path)[1].lower()
        
        if test_ext not in ['.xlsx', '.xls'] or doc_ext not in ['.xlsx', '.xls']:
            raise ValueError("Cả hai file phải là file Excel (.xlsx hoặc .xls)")
        
        if progress_callback:
            progress_callback("Đang đọc file mocks...")
        
        # Load the mocks file (source)
        source_wb = load_workbook(test_file_path, data_only=True)
        
        # Find the アクション一覧 sheet in mocks file
        sheet_name = "アクション一覧"
        if sheet_name not in source_wb.sheetnames:
            raise ValueError(f"Không tìm thấy sheet '{sheet_name}' trong file mocks")
        
        source_ws = source_wb[sheet_name]
        
        if progress_callback:
            progress_callback("Đang tìm cột アクションNo....")
        
        # Find the "アクションNo." header cell
        header_cell = None
        header_row = None
        header_col = None
        
        for row in range(1, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                cell = source_ws.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if cell_text == "アクションNo.":
                        header_cell = cell
                        header_row = row
                        header_col = col
                        break
            if header_cell:
                break
        
        if not header_cell:
            raise ValueError("Không tìm thấy cột 'アクションNo.' trong sheet mocks")
        
        if progress_callback:
            progress_callback("Đang đọc dữ liệu từ cột アクションNo....")
        
        # Get all data below the header in the same column
        data_to_copy = []
        
        for row in range(header_row + 1, source_ws.max_row + 1):
            cell = source_ws.cell(row=row, column=header_col)
            if cell.value is not None:  # Only copy non-empty cells
                data_to_copy.append(cell)
            # Continue even if we encounter empty cells to get all data
        
        # Remove trailing empty cells if any
        while data_to_copy and data_to_copy[-1].value is None:
            data_to_copy.pop()
        
        if not data_to_copy:
            raise ValueError("Không có dữ liệu nào trong cột 'アクションNo.' để copy")
        
        if progress_callback:
            progress_callback("Đang mở file unit test...")
        
        # Load the unit test file (destination)
        dest_wb = load_workbook(doc_file_path)
        
        # Find or create the アクション一覧 sheet in unit test file
        if sheet_name not in dest_wb.sheetnames:
            dest_ws = dest_wb.create_sheet(sheet_name)
        else:
            dest_ws = dest_wb[sheet_name]
        
        if progress_callback:
            progress_callback("Đang tìm cột アクションNo. trong file unit test...")
        
        # Find the "アクションNo." header cell in destination
        dest_header_cell = None
        dest_header_row = None
        dest_header_col = None
        
        for row in range(1, dest_ws.max_row + 1):
            for col in range(1, dest_ws.max_column + 1):
                cell = dest_ws.cell(row=row, column=col)
                if cell.value and str(cell.value).strip() == "アクションNo.":
                    dest_header_cell = cell
                    dest_header_row = row
                    dest_header_col = col
                    break
            if dest_header_cell:
                break
        
        if not dest_header_cell:
            # If header doesn't exist, create it at A1
            dest_header_row = 1
            dest_header_col = 1
            dest_ws.cell(row=dest_header_row, column=dest_header_col, value="アクションNo.")
        
        if progress_callback:
            progress_callback("Đang sao chép dữ liệu...")
        
        # Paste data below the header in destination
        paste_start_row = dest_header_row + 1
        
        for idx, source_cell in enumerate(data_to_copy):
            dest_cell = dest_ws.cell(
                row=paste_start_row + idx,
                column=dest_header_col
            )
            
            # Copy cell value
            dest_cell.value = source_cell.value
            
            # Copy cell formatting if source cell has formatting
            if source_cell.has_style:
                current_font = copy(source_cell.font)
                current_font.color = "000000"  # Set font color to black
                dest_cell.font = current_font
                dest_cell.border = copy(source_cell.border)
                dest_cell.fill = copy(source_cell.fill)
                dest_cell.number_format = source_cell.number_format
                dest_cell.protection = copy(source_cell.protection)
                dest_cell.alignment = copy(source_cell.alignment)
        
        print(f"Completed copying {len(data_to_copy)} items to アクションNo. column (no merging applied)")

        # Also copy data from "アクション" column to "項目" and "操作" columns
        if progress_callback:
            progress_callback("Đang tìm cột アクション...")
        
        # Find the "アクション" header cell in source
        action_header_cell = None
        action_header_row = None
        action_header_col = None
        
        for row in range(1, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                cell = source_ws.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if cell_text == "アクション":
                        action_header_cell = cell
                        action_header_row = row
                        action_header_col = col

                        break
            if action_header_cell:
                break
        
        if action_header_cell:
            # Get data from アクション column
            action_data = []
            for row in range(action_header_row + 1, source_ws.max_row + 1):
                cell = source_ws.cell(row=row, column=action_header_col)
                if cell.value is not None:

                    action_data.append(cell)
            
            if action_data:
                if progress_callback:
                    progress_callback("Đang copy dữ liệu アクション sang 項目 và 操作...")
                
                # Find 項目 and 操作 columns in destination
                target_columns = {"項目": None, "操作": None}
                
                for row in range(1, dest_ws.max_row + 1):
                    for col in range(1, dest_ws.max_column + 1):
                        cell = dest_ws.cell(row=row, column=col)
                        if cell.value:
                            cell_text = str(cell.value).strip()
                            if cell_text in target_columns and target_columns[cell_text] is None:
                                target_columns[cell_text] = {"row": row, "col": col}
                                print(f"Found target column '{cell_text}' at row {row}, column {col}")
                
                # Create headers if they don't exist
                next_col = dest_ws.max_column + 1
                for header_name, header_info in target_columns.items():
                    if header_info is None:
                        dest_ws.cell(row=1, column=next_col, value=header_name)
                        target_columns[header_name] = {"row": 1, "col": next_col}
                        print(f"Created header '{header_name}' at column {next_col}")
                        next_col += 1
                
                # Copy action data to both target columns
                for target_name, target_info in target_columns.items():
                    target_col = target_info["col"]
                    target_start_row = target_info["row"] + 1
                    
                    for idx, source_cell in enumerate(action_data):
                        dest_cell = dest_ws.cell(
                            row=target_start_row + idx,
                            column=target_col
                        )
                        
                        # Copy cell value
                        dest_cell.value = source_cell.value
                        print(f"Copied '{source_cell.value}' to {target_name} column at row {target_start_row + idx}")
                        
                        # Copy cell formatting
                        if source_cell.has_style:
                            dest_cell.font = copy(source_cell.font)
                            dest_cell.border = copy(source_cell.border)
                            dest_cell.fill = copy(source_cell.fill)
                            dest_cell.number_format = source_cell.number_format
                            dest_cell.protection = copy(source_cell.protection)
                            dest_cell.alignment = copy(source_cell.alignment)
                    
                    print(f"Completed copying {len(action_data)} items to {target_name} column (no merging applied)")
            else:
                print("No data found in アクション column")
        else:
            print("アクション column not found")

        # Also copy data from "処理条件" column to "処理条件①" column
        if progress_callback:
            progress_callback("Đang tìm cột 処理条件...")
        
        # Find the "処理条件" header cell in source
        condition_header_cell = None
        condition_header_row = None
        condition_header_col = None
        
        for row in range(1, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                cell = source_ws.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if cell_text == "処理条件":
                        condition_header_cell = cell
                        condition_header_row = row
                        condition_header_col = col
                        print(f"Condition header found at row {condition_header_row}, column {condition_header_col}")
                        break
            if condition_header_cell:
                break
        
        if condition_header_cell:
            # Get data from 処理条件 column
            condition_data = []
            for row in range(condition_header_row + 1, source_ws.max_row + 1):
                cell = source_ws.cell(row=row, column=condition_header_col)
                if cell.value is not None:
                    print(f"Found condition data at row {row}: '{cell.value}'")
                    condition_data.append(cell)
            
            if condition_data:
                if progress_callback:
                    progress_callback("Đang copy dữ liệu 処理条件 sang 処理条件①...")
                
                # Find or create 処理条件① column in destination
                target_condition_col = None
                target_condition_row = None
                
                for row in range(1, dest_ws.max_row + 1):
                    for col in range(1, dest_ws.max_column + 1):
                        cell = dest_ws.cell(row=row, column=col)
                        if cell.value:
                            cell_text = str(cell.value).strip()
                            if cell_text == "処理条件①":
                                target_condition_col = col
                                target_condition_row = row
                                print(f"Found target column '処理条件①' at row {row}, column {col}")
                                break
                    if target_condition_col:
                        break
                
                # Create header if it doesn't exist
                if target_condition_col is None:
                    target_condition_col = dest_ws.max_column + 1
                    target_condition_row = 1
                    dest_ws.cell(row=target_condition_row, column=target_condition_col, value="処理条件①")
                    print(f"Created header '処理条件①' at column {target_condition_col}")
                
                # Copy condition data to target column
                target_start_row = target_condition_row + 1
                
                for idx, source_cell in enumerate(condition_data):
                    dest_cell = dest_ws.cell(
                        row=target_start_row + idx,
                        column=target_condition_col
                    )
                    
                    # Copy cell value
                    dest_cell.value = source_cell.value
                    print(f"Copied '{source_cell.value}' to 処理条件① column at row {target_start_row + idx}")
                    
                    # Copy cell formatting
                    if source_cell.has_style:
                        dest_cell.font = copy(source_cell.font)
                        dest_cell.border = copy(source_cell.border)
                        dest_cell.fill = copy(source_cell.fill)
                        dest_cell.number_format = source_cell.number_format
                        dest_cell.protection = copy(source_cell.protection)
                        dest_cell.alignment = copy(source_cell.alignment)
            else:
                print("No data found in 処理条件 column")
        else:
            print("処理条件 column not found")

        # Also copy data from "API URL" column to "WEBAPI" column
        if progress_callback:
            progress_callback("Đang tìm cột API URL...")
        
        # Find the "API URL" header cell in source
        api_header_cell = None
        api_header_row = None
        api_header_col = None
        
        for row in range(1, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                cell = source_ws.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if cell_text == "API URL":
                        api_header_cell = cell
                        api_header_row = row
                        api_header_col = col
                        print(f"API URL header found at row {api_header_row}, column {api_header_col}")
                        break
            if api_header_cell:
                break
        
        if api_header_cell:
            # Get data from API URL column
            api_data = []
            for row in range(api_header_row + 1, source_ws.max_row + 1):
                cell = source_ws.cell(row=row, column=api_header_col)
                if cell.value is not None:
                    print(f"Found API URL data at row {row}: '{cell.value}'")
                    api_data.append(cell)
            
            if api_data:
                if progress_callback:
                    progress_callback("Đang copy dữ liệu API URL sang WEBAPI...")
                
                # Find or create WEBAPI column in destination
                target_webapi_col = None
                target_webapi_row = None
                
                for row in range(1, dest_ws.max_row + 1):
                    for col in range(1, dest_ws.max_column + 1):
                        cell = dest_ws.cell(row=row, column=col)
                        if cell.value:
                            cell_text = str(cell.value).strip()
                            if cell_text == "WEBAPI":
                                target_webapi_col = col
                                target_webapi_row = row
                                print(f"Found target column 'WEBAPI' at row {row}, column {col}")
                                break
                    if target_webapi_col:
                        break
                
                # Create header if it doesn't exist
                if target_webapi_col is None:
                    target_webapi_col = dest_ws.max_column + 1
                    target_webapi_row = 1
                    dest_ws.cell(row=target_webapi_row, column=target_webapi_col, value="WEBAPI")
                    print(f"Created header 'WEBAPI' at column {target_webapi_col}")
                
                # Copy API data to WEBAPI column (without merging consecutive cells)
                target_start_row = target_webapi_row + 1
                
                for idx, source_cell in enumerate(api_data):
                    dest_cell = dest_ws.cell(
                        row=target_start_row + idx,
                        column=target_webapi_col
                    )
                    
                    # Copy cell value
                    dest_cell.value = source_cell.value
                    print(f"Copied '{source_cell.value}' to WEBAPI column at row {target_start_row + idx}")
                    
                    # Copy cell formatting
                    if source_cell.has_style:
                        dest_cell.font = copy(source_cell.font)
                        dest_cell.border = copy(source_cell.border)
                        dest_cell.fill = copy(source_cell.fill)
                        dest_cell.number_format = source_cell.number_format
                        dest_cell.protection = copy(source_cell.protection)
                        dest_cell.alignment = copy(source_cell.alignment)
            else:
                print("No data found in API URL column")
        else:
            print("API URL column not found")

        # Also copy data from column containing "処理No." to "No." column
        if progress_callback:
            progress_callback("Đang tìm cột chứa 処理No....")
        
        # Find the column containing "処理No." header cell in source
        shori_no_header_cell = None
        shori_no_header_row = None
        shori_no_header_col = None
        
        for row in range(1, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                cell = source_ws.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if "処理No." in cell_text:  # Contains "処理No." rather than exact match
                        shori_no_header_cell = cell
                        shori_no_header_row = row
                        shori_no_header_col = col
                        print(f"Column containing '処理No.' found at row {shori_no_header_row}, column {shori_no_header_col}: '{cell_text}'")
                        break
            if shori_no_header_cell:
                break
        
        if shori_no_header_cell:
            # Get data from 処理No. column
            shori_no_data = []
            for row in range(shori_no_header_row + 1, source_ws.max_row + 1):
                cell = source_ws.cell(row=row, column=shori_no_header_col)
                if cell.value is not None:
                    print(f"Found 処理No. data at row {row}: '{cell.value}'")
                    shori_no_data.append(cell)
            
            if shori_no_data:
                if progress_callback:
                    progress_callback("Đang copy dữ liệu 処理No. sang No....")
                
                # Find or create No. column in destination
                target_no_col = None
                target_no_row = None
                
                for row in range(1, dest_ws.max_row + 1):
                    for col in range(1, dest_ws.max_column + 1):
                        cell = dest_ws.cell(row=row, column=col)
                        if cell.value:
                            cell_text = str(cell.value).strip()
                            if cell_text == "No.":
                                target_no_col = col
                                target_no_row = row
                                print(f"Found target column 'No.' at row {row}, column {col}")
                                break
                    if target_no_col:
                        break
                
                # Create header if it doesn't exist
                if target_no_col is None:
                    target_no_col = dest_ws.max_column + 1
                    target_no_row = 1
                    dest_ws.cell(row=target_no_row, column=target_no_col, value="No.")
                    print(f"Created header 'No.' at column {target_no_col}")
                
                # Copy 処理No. data to No. column (without merging consecutive cells)
                target_start_row = target_no_row + 1
                
                for idx, source_cell in enumerate(shori_no_data):
                    dest_cell = dest_ws.cell(
                        row=target_start_row + idx,
                        column=target_no_col
                    )
                    
                    # Copy cell value
                    dest_cell.value = source_cell.value
                    print(f"Copied '{source_cell.value}' to No. column at row {target_start_row + idx}")
                    
                    # Copy cell formatting
                    if source_cell.has_style:
                        dest_cell.font = copy(source_cell.font)
                        dest_cell.border = copy(source_cell.border)
                        dest_cell.fill = copy(source_cell.fill)
                        dest_cell.number_format = source_cell.number_format
                        dest_cell.protection = copy(source_cell.protection)
                        dest_cell.alignment = copy(source_cell.alignment)
                
                # Format No. column data - keep only the last character
                if progress_callback:
                    progress_callback("Đang format cột No. - chỉ giữ ký tự cuối...")
                
                print("Formatting No. column data to keep only last character...")
                for idx, source_cell in enumerate(shori_no_data):
                    dest_cell = dest_ws.cell(
                        row=target_start_row + idx,
                        column=target_no_col
                    )
                    
                    # Get the current value and keep only the last character
                    current_value = str(dest_cell.value).strip() if dest_cell.value is not None else ""
                    if current_value:
                        last_char = current_value[-1]  # Get the last character
                        
                        # Check if it's a number and format appropriately
                        if last_char.isdigit():
                            dest_cell.value = int(last_char)  # Set as integer to avoid text format warning
                            dest_cell.number_format = '0'  # Format as number
                        else:
                            dest_cell.value = last_char  # Set as text
                            dest_cell.number_format = '@'  # Format as text
                        
                        # Set font color to black
                        if dest_cell.font:
                            from openpyxl.styles import Font
                            current_font = copy(dest_cell.font)
                            current_font.color = "000000"  # Black color
                            dest_cell.font = current_font
                        
                        print(f"Formatted No. at row {target_start_row + idx}: '{current_value}' → '{last_char}' ({'number' if last_char.isdigit() else 'text'})")
                    else:
                        print(f"No data to format at row {target_start_row + idx}")
            else:
                print("No data found in 処理No. column")
        else:
            print("Column containing '処理No.' not found")

        # Copy combined data from "処理概要" and "入力パラメータ" columns to "想定結果" column
        if progress_callback:
            progress_callback("Đang tìm cột 処理概要 và 入力パラメータ...")
        
        # Find the "処理概要" and "入力パラメータ" header cells in source
        shori_gaiyo_header_cell = None
        shori_gaiyo_header_row = None
        shori_gaiyo_header_col = None
        
        nyuryoku_param_header_cell = None
        nyuryoku_param_header_row = None
        nyuryoku_param_header_col = None
        
        # Find 処理概要 column
        for row in range(1, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                cell = source_ws.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if cell_text == "処理概要":
                        shori_gaiyo_header_cell = cell
                        shori_gaiyo_header_row = row
                        shori_gaiyo_header_col = col
                        print(f"処理概要 header found at row {shori_gaiyo_header_row}, column {shori_gaiyo_header_col}")
                        break
            if shori_gaiyo_header_cell:
                break
        
        # Find 入力パラメータ column
        for row in range(1, source_ws.max_row + 1):
            for col in range(1, source_ws.max_column + 1):
                cell = source_ws.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if cell_text == "入力パラメータ":
                        nyuryoku_param_header_cell = cell
                        nyuryoku_param_header_row = row
                        nyuryoku_param_header_col = col
                        print(f"入力パラメータ header found at row {nyuryoku_param_header_row}, column {nyuryoku_param_header_col}")
                        break
            if nyuryoku_param_header_cell:
                break
        
        if shori_gaiyo_header_cell and nyuryoku_param_header_cell:
            if progress_callback:
                progress_callback("Đang kết hợp dữ liệu từ 2 cột...")
            
            # Get combined data from both columns
            combined_data = []
            max_data_row = max(source_ws.max_row, 
                              shori_gaiyo_header_row if shori_gaiyo_header_cell else 0,
                              nyuryoku_param_header_row if nyuryoku_param_header_cell else 0)
            
            # Determine the start row (take the maximum of both headers + 1)
            start_row = max(shori_gaiyo_header_row + 1, nyuryoku_param_header_row + 1)
            
            for row in range(start_row, max_data_row + 1):
                shori_gaiyo_cell = source_ws.cell(row=row, column=shori_gaiyo_header_col)
                nyuryoku_param_cell = source_ws.cell(row=row, column=nyuryoku_param_header_col)
                
                # Get values from both cells
                shori_gaiyo_value = str(shori_gaiyo_cell.value).strip() if shori_gaiyo_cell.value is not None else ""
                nyuryoku_param_value = str(nyuryoku_param_cell.value).strip() if nyuryoku_param_cell.value is not None else ""
                
                # Skip if 処理概要 is empty
                if not shori_gaiyo_value:
                    continue
                
                # Only combine if 入力パラメータ has meaningful data (not empty and not "-")
                if nyuryoku_param_value and nyuryoku_param_value != "-":
                    # Combine values with newline separator
                    combined_value = f"{shori_gaiyo_value}\n{nyuryoku_param_value}"
                else:
                    # Only use 処理概要 value
                    combined_value = shori_gaiyo_value
                
                if nyuryoku_param_value and nyuryoku_param_value != "-":
                    print(f"Combined data at row {row}: '{combined_value.replace(chr(10), '\\n')}' (処理概要 + 入力パラメータ)")
                else:
                    print(f"Single data at row {row}: '{combined_value}' (chỉ 処理概要, 入力パラメータ: '{nyuryoku_param_value or 'rỗng'}')")
                
                # Store the combined data with source cell for formatting reference (use the first non-empty cell)
                source_cell_for_format = shori_gaiyo_cell if shori_gaiyo_cell.value is not None else nyuryoku_param_cell
                combined_data.append({
                    'value': combined_value,
                    'source_cell': source_cell_for_format,
                    'row': row
                })
            
            if combined_data:
                if progress_callback:
                    progress_callback("Đang copy dữ liệu kết hợp sang 想定結果...")
                
                # Find or create 想定結果 column in destination
                target_soutei_col = None
                target_soutei_row = None
                
                for row in range(1, dest_ws.max_row + 1):
                    for col in range(1, dest_ws.max_column + 1):
                        cell = dest_ws.cell(row=row, column=col)
                        if cell.value:
                            cell_text = str(cell.value).strip()
                            if cell_text == "想定結果":
                                target_soutei_col = col
                                target_soutei_row = row
                                print(f"Found target column '想定結果' at row {row}, column {col}")
                                break
                    if target_soutei_col:
                        break
                
                # Create header if it doesn't exist
                if target_soutei_col is None:
                    target_soutei_col = dest_ws.max_column + 1
                    target_soutei_row = 1
                    dest_ws.cell(row=target_soutei_row, column=target_soutei_col, value="想定結果")
                    print(f"Created header '想定結果' at column {target_soutei_col}")
                
                # Copy combined data to 想定結果 column
                target_start_row = target_soutei_row + 1
                
                for idx, data_item in enumerate(combined_data):
                    dest_cell = dest_ws.cell(
                        row=target_start_row + idx,
                        column=target_soutei_col
                    )
                    
                    # Copy combined value
                    dest_cell.value = data_item['value']
                    print(f"Copied combined value to 想定結果 column at row {target_start_row + idx}")
                    
                    # Copy cell formatting from source
                    source_cell = data_item['source_cell']
                    if source_cell.has_style:
                        dest_cell.font = copy(source_cell.font)
                        dest_cell.border = copy(source_cell.border)
                        dest_cell.fill = copy(source_cell.fill)
                        dest_cell.number_format = source_cell.number_format
                        dest_cell.protection = copy(source_cell.protection)
                        dest_cell.alignment = copy(source_cell.alignment)
            else:
                print("No data found in 処理概要 and 入力パラメータ columns")
        else:
            if not shori_gaiyo_header_cell:
                print("処理概要 column not found")
            if not nyuryoku_param_header_cell:
                print("入力パラメータ column not found")

        # Fill existing additional condition columns (処理条件② through 処理条件⑥) with "-"
        if progress_callback:
            progress_callback("Đang điền dữ liệu vào các cột 処理条件② đến 処理条件⑥...")
        
        # List of additional condition columns to fill
        additional_condition_columns = ["処理条件②", "処理条件③", "処理条件④", "処理条件⑤", "処理条件⑥"]
        columns_filled = []
        
        # Determine the row count based on 処理条件① column
        condition_row_count = 0
        if 'condition_data' in locals() and condition_data:
            condition_row_count = len(condition_data)
            print(f"Using row count from 処理条件①: {condition_row_count}")
        else:
            # If no 処理条件① data, use the same count as アクションNo. data
            condition_row_count = len(data_to_copy)
            print(f"Using row count from アクションNo.: {condition_row_count}")
        
        if condition_row_count > 0:
            # Create thin border for these columns
            thin_border = Side(style='thin', color='000000')
            cell_border = Border(
                top=thin_border,
                bottom=thin_border,
                left=thin_border,
                right=thin_border
            )
            
            for col_name in additional_condition_columns:
                # Find existing column (don't create new ones)
                target_col = None
                target_row = None
                
                # Search for existing column
                for row in range(1, dest_ws.max_row + 1):
                    for col in range(1, dest_ws.max_column + 1):
                        cell = dest_ws.cell(row=row, column=col)
                        if cell.value and str(cell.value).strip() == col_name:
                            target_col = col
                            target_row = row
                            print(f"Found existing column '{col_name}' at row {row}, column {col}")
                            break
                    if target_col:
                        break
                
                # Only fill if column exists
                if target_col is not None:
                    # Fill the column with "-" values and apply borders
                    start_data_row = target_row + 1
                    for i in range(condition_row_count):
                        cell_row = start_data_row + i
                        cell = dest_ws.cell(row=cell_row, column=target_col)
                        
                        # Set value to "-"
                        cell.value = "-"
                        
                        # Apply border
                        cell.border = cell_border
                        
                        print(f"Set '{col_name}' cell at row {cell_row} to '-' with border")
                    
                    columns_filled.append(col_name)
                else:
                    print(f"Column '{col_name}' not found in unit test file")
        else:
            print("No data found to determine row count for additional condition columns")

        # Add borders to existing columns 実施者, 実施日, 結果 (without filling data)
        if progress_callback:
            progress_callback("Đang thêm border cho các cột 実施者, 実施日, 結果...")
        
        # List of columns to add borders to (without data)
        border_only_columns = ["実施者", "実施日", "結果"]
        columns_bordered = []
        
        # Use the same row count as previous operations
        border_row_count = 0
        if 'condition_data' in locals() and condition_data:
            border_row_count = len(condition_data)
            print(f"Using row count from 処理条件①: {border_row_count}")
        else:
            # If no 処理条件① data, use the same count as アクションNo. data
            border_row_count = len(data_to_copy)
            print(f"Using row count from アクションNo.: {border_row_count}")
        
        if border_row_count > 0:
            # Create thin border for these columns
            thin_border = Side(style='thin', color='000000')
            cell_border = Border(
                top=thin_border,
                bottom=thin_border,
                left=thin_border,
                right=thin_border
            )
            
            for col_name in border_only_columns:
                # Find existing column (don't create new ones)
                target_col = None
                target_row = None
                
                # Search for existing column
                for row in range(1, dest_ws.max_row + 1):
                    for col in range(1, dest_ws.max_column + 1):
                        cell = dest_ws.cell(row=row, column=col)
                        if cell.value and str(cell.value).strip() == col_name:
                            target_col = col
                            target_row = row
                            print(f"Found existing column '{col_name}' at row {row}, column {col}")
                            break
                    if target_col:
                        break
                
                # Only add borders if column exists
                if target_col is not None:
                    # Add borders to the column cells (without filling data)
                    start_data_row = target_row + 1
                    for i in range(border_row_count):
                        cell_row = start_data_row + i
                        cell = dest_ws.cell(row=cell_row, column=target_col)
                        
                        # Apply border only (don't change the value)
                        cell.border = cell_border
                        
                        print(f"Added border to '{col_name}' cell at row {cell_row}")
                    
                    columns_bordered.append(col_name)
                else:
                    print(f"Column '{col_name}' not found in unit test file")
        else:
            print("No data found to determine row count for border-only columns")

        # Process row merging for No. and 想定結果 columns
        if progress_callback:
            progress_callback("Đang gộp hàng có cùng điều kiện...")
        
        # Find columns for merging
        no_col = None
        action_no_col = None
        koumoku_col = None  # 項目
        sousa_col = None    # 操作
        soutei_col = None
        
        for row in range(1, dest_ws.max_row + 1):
            for col in range(1, dest_ws.max_column + 1):
                cell = dest_ws.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).strip()
                    if cell_text == "No." and no_col is None:
                        no_col = col
                    elif cell_text == "アクションNo." and action_no_col is None:
                        action_no_col = col
                    elif cell_text == "項目" and koumoku_col is None:
                        koumoku_col = col
                    elif cell_text == "操作" and sousa_col is None:
                        sousa_col = col
                    elif cell_text == "想定結果" and soutei_col is None:
                        soutei_col = col
        
        rows_processed = 0
        rows_deleted = 0
        
        if no_col is not None and action_no_col is not None and soutei_col is not None:
            print(f"DEBUG: Found No. column at {no_col}, アクションNo. column at {action_no_col}, 項目 column at {koumoku_col}, 操作 column at {sousa_col}, 想定結果 column at {soutei_col}")
            print(f"DEBUG: Processing rows from 2 to {dest_ws.max_row}")
            
            # Get all data rows (skip header row)
            data_rows = []
            for row in range(2, dest_ws.max_row + 1):  # Start from row 2 to skip headers
                no_cell = dest_ws.cell(row=row, column=no_col)
                action_no_cell = dest_ws.cell(row=row, column=action_no_col)
                soutei_cell = dest_ws.cell(row=row, column=soutei_col)
                
                # Only process rows that have data in all three columns
                if no_cell.value is not None and action_no_cell.value is not None and soutei_cell.value is not None:
                    no_value = str(no_cell.value).strip()
                    action_no_value = str(action_no_cell.value).strip()
                    soutei_value = str(soutei_cell.value).strip()
                    
                    # Split 想定結果 by newlines to get first line and second line
                    soutei_lines = soutei_value.split('\n')
                    first_line = soutei_lines[0].strip() if soutei_lines else ""
                    second_line = soutei_lines[1].strip() if len(soutei_lines) > 1 else ""
                    
                    print(f"DEBUG: Row {row} - No='{no_value}', ActionNo='{action_no_value}', First line='{first_line[:50]}...'")
                    
                    data_rows.append({
                        'row': row,
                        'no_value': no_value,
                        'action_no_value': action_no_value,
                        'soutei_value': soutei_value,
                        'soutei_first_line': first_line,
                        'soutei_second_line': second_line,
                        'soutei_lines': soutei_lines
                    })
                else:
                    print(f"DEBUG: Row {row} skipped - No='{no_cell.value}', ActionNo='{action_no_cell.value}', 想定結果='{soutei_cell.value}'")
            
            # Group rows by No., ActionNo. and first line of 想定結果
            groups = {}
            for data_row in data_rows:
                key = f"{data_row['no_value']}||{data_row['action_no_value']}||{data_row['soutei_first_line']}"
                if key not in groups:
                    groups[key] = []
                groups[key].append(data_row)
            
            # Debug: Print all groups found
            print(f"DEBUG: Found {len(groups)} groups:")
            for key, group_rows in groups.items():
                print(f"  Group '{key}': {len(group_rows)} rows")
                for row_data in group_rows:
                    print(f"    Row {row_data['row']}: No='{row_data['no_value']}', ActionNo='{row_data['action_no_value']}', First line='{row_data['soutei_first_line']}'")
            
            # Process each group that has multiple rows
            rows_to_delete = []
            
            for key, group_rows in groups.items():
                if len(group_rows) > 1:  # Only process groups with multiple rows
                    # Sort by row number to ensure consistent processing
                    group_rows.sort(key=lambda x: x['row'])
                    
                    # First row will be the target row
                    target_row = group_rows[0]
                    
                    # Collect all second lines from rows 2 onwards in this group
                    additional_lines = []
                    for row_data in group_rows[1:]:  # Skip first row
                        print(f"DEBUG: Processing row {row_data['row']} (No='{row_data['no_value']}', ActionNo='{row_data['action_no_value']}') for deletion")
                        if row_data['soutei_second_line']:
                            additional_lines.append(row_data['soutei_second_line'])
                        # Mark this row for deletion
                        rows_to_delete.append(row_data['row'])
                    
                    # Update target row's 想定結果 with combined data
                    if additional_lines:
                        # Combine: first_line + original_second_line + additional_lines
                        all_lines = [target_row['soutei_first_line']]
                        if target_row['soutei_second_line']:
                            all_lines.append(target_row['soutei_second_line'])
                        all_lines.extend(additional_lines)
                        
                        new_soutei_value = '\n'.join(all_lines)
                        
                        # Update the target cell
                        target_cell = dest_ws.cell(row=target_row['row'], column=soutei_col)
                        target_cell.value = new_soutei_value
                    
                    rows_processed += 1
            
            # Delete marked rows in reverse order (from bottom to top to avoid index shifting)
            if rows_to_delete:
                print(f"DEBUG: Rows marked for deletion: {sorted(rows_to_delete)}")
                rows_to_delete.sort(reverse=True)  # Sort in descending order
                
                for row_num in rows_to_delete:
                    # Get No. and ActionNo. values before deletion for debug
                    no_cell = dest_ws.cell(row=row_num, column=no_col)
                    action_no_cell = dest_ws.cell(row=row_num, column=action_no_col)
                    no_value = str(no_cell.value).strip() if no_cell.value else "None"
                    action_no_value = str(action_no_cell.value).strip() if action_no_cell.value else "None"
                    print(f"DEBUG: Deleting row {row_num} (No='{no_value}', ActionNo='{action_no_value}')")
                    dest_ws.delete_rows(row_num)
                
                rows_deleted = len(rows_to_delete)
            else:
                print("DEBUG: No rows marked for deletion")
        else:
            print(f"DEBUG: Required columns not found - No.: {no_col}, ActionNo.: {action_no_col}, 想定結果: {soutei_col}")
        
        # Merge cells in アクションNo. column for rows with same ActionNo
        if progress_callback:
            progress_callback("Đang merge cột アクションNo....")
        
        action_merges = 0
        
        if action_no_col is not None:
            print(f"DEBUG: Processing ActionNo column merging at column {action_no_col}")
            
            # Get all data rows after previous processing
            current_max_row = dest_ws.max_row
            
            # Collect ActionNo values and their row ranges
            action_groups = []
            current_group = None
            
            for row in range(2, current_max_row + 1):  # Start from row 2 to skip headers
                action_cell = dest_ws.cell(row=row, column=action_no_col)
                action_value = str(action_cell.value).strip() if action_cell.value else ""
                
                if action_value:  # Only process non-empty cells
                    if current_group is None or current_group['value'] != action_value:
                        # Start new group
                        if current_group is not None:
                            action_groups.append(current_group)
                        current_group = {
                            'value': action_value,
                            'start_row': row,
                            'end_row': row
                        }
                    else:
                        # Extend current group
                        current_group['end_row'] = row
            
            # Add the last group
            if current_group is not None:
                action_groups.append(current_group)
            
            # Process merging for groups with multiple rows
            for group in action_groups:
                if group['end_row'] > group['start_row']:  # Multiple rows
                    start_row = group['start_row']
                    end_row = group['end_row']
                    print(f"DEBUG: Merging ActionNo '{group['value']}' from row {start_row} to {end_row}")
                    
                    # Merge the cells in ActionNo column
                    dest_ws.merge_cells(start_row=start_row, start_column=action_no_col, 
                                       end_row=end_row, end_column=action_no_col)
                    
                    action_merges += 1
        
        print(f"DEBUG: Completed ActionNo column merging - {action_merges} groups merged")
        
        # Merge cells in 項目 column for rows with same 項目
        if progress_callback:
            progress_callback("Đang merge cột 項目...")
        
        koumoku_merges = 0
        
        if koumoku_col is not None:
            print(f"DEBUG: Processing 項目 column merging at column {koumoku_col}")
            
            # Get all data rows after previous processing
            current_max_row = dest_ws.max_row
            
            # Collect 項目 values and their row ranges
            koumoku_groups = []
            current_group = None
            
            for row in range(2, current_max_row + 1):  # Start from row 2 to skip headers
                koumoku_cell = dest_ws.cell(row=row, column=koumoku_col)
                koumoku_value = str(koumoku_cell.value).strip() if koumoku_cell.value else ""
                
                if koumoku_value:  # Only process non-empty cells
                    if current_group is None or current_group['value'] != koumoku_value:
                        # Start new group
                        if current_group is not None:
                            koumoku_groups.append(current_group)
                        current_group = {
                            'value': koumoku_value,
                            'start_row': row,
                            'end_row': row
                        }
                    else:
                        # Extend current group
                        current_group['end_row'] = row
            
            # Add the last group
            if current_group is not None:
                koumoku_groups.append(current_group)
            
            # Process merging for groups with multiple rows
            for group in koumoku_groups:
                if group['end_row'] > group['start_row']:  # Multiple rows
                    start_row = group['start_row']
                    end_row = group['end_row']
                    print(f"DEBUG: Merging 項目 '{group['value']}' from row {start_row} to {end_row}")
                    
                    # Merge the cells in 項目 column
                    dest_ws.merge_cells(start_row=start_row, start_column=koumoku_col, 
                                       end_row=end_row, end_column=koumoku_col)
                    
                    koumoku_merges += 1
        
        print(f"DEBUG: Completed 項目 column merging - {koumoku_merges} groups merged")
        
        # Merge cells in 操作 column for rows with same 操作
        if progress_callback:
            progress_callback("Đang merge cột 操作...")
        
        sousa_merges = 0
        
        if sousa_col is not None:
            print(f"DEBUG: Processing 操作 column merging at column {sousa_col}")
            
            # Get all data rows after previous processing
            current_max_row = dest_ws.max_row
            
            # Collect 操作 values and their row ranges
            sousa_groups = []
            current_group = None
            
            for row in range(2, current_max_row + 1):  # Start from row 2 to skip headers
                sousa_cell = dest_ws.cell(row=row, column=sousa_col)
                sousa_value = str(sousa_cell.value).strip() if sousa_cell.value else ""
                
                if sousa_value:  # Only process non-empty cells
                    if current_group is None or current_group['value'] != sousa_value:
                        # Start new group
                        if current_group is not None:
                            sousa_groups.append(current_group)
                        current_group = {
                            'value': sousa_value,
                            'start_row': row,
                            'end_row': row
                        }
                    else:
                        # Extend current group
                        current_group['end_row'] = row
            
            # Add the last group
            if current_group is not None:
                sousa_groups.append(current_group)
            
            # Process merging for groups with multiple rows
            for group in sousa_groups:
                if group['end_row'] > group['start_row']:  # Multiple rows
                    start_row = group['start_row']
                    end_row = group['end_row']
                    print(f"DEBUG: Merging 操作 '{group['value']}' from row {start_row} to {end_row}")
                    
                    # Merge the cells in 操作 column
                    dest_ws.merge_cells(start_row=start_row, start_column=sousa_col, 
                                       end_row=end_row, end_column=sousa_col)
                    
                    sousa_merges += 1
        
        print(f"DEBUG: Completed 操作 column merging - {sousa_merges} groups merged")
            
        if progress_callback:
            progress_callback("Đang lưu file...")
        
        # Save the destination file
        dest_wb.save(doc_file_path)
        
        # Close workbooks
        source_wb.close()
        dest_wb.close()
        
        result = {
            "test_file": os.path.basename(test_file_path),
            "doc_file": os.path.basename(doc_file_path),
            "sheet_name": sheet_name,
            "header_found_at": f"Row {header_row}, Col {header_col}",
            "data_copied": len(data_to_copy),
            "paste_location": f"Row {paste_start_row}, Col {dest_header_col}",
            "dest_header_created": dest_header_cell is None,
            "action_column_found": action_header_cell is not None,
            "action_data_copied": len(action_data) if 'action_data' in locals() and action_data else 0,
            "target_columns_created": "項目, 操作" if action_header_cell else "None",
            "condition_column_found": condition_header_cell is not None,
            "condition_data_copied": len(condition_data) if 'condition_data' in locals() and condition_data else 0,
            "condition_target_created": "処理条件①" if condition_header_cell else "None",
            "api_column_found": api_header_cell is not None,
            "api_data_copied": len(api_data) if 'api_data' in locals() and api_data else 0,
            "webapi_target_created": "WEBAPI" if api_header_cell else "None",
            "shori_no_column_found": shori_no_header_cell is not None,
            "shori_no_data_copied": len(shori_no_data) if 'shori_no_data' in locals() and shori_no_data else 0,
            "no_target_created": "No." if shori_no_header_cell else "None",
            "no_data_formatted": len(shori_no_data) if 'shori_no_data' in locals() and shori_no_data else 0,
            "combined_columns_found": shori_gaiyo_header_cell is not None and nyuryoku_param_header_cell is not None,
            "combined_data_copied": len(combined_data) if 'combined_data' in locals() and combined_data else 0,
            "soutei_target_created": "想定結果" if 'shori_gaiyo_header_cell' in locals() and shori_gaiyo_header_cell and 'nyuryoku_param_header_cell' in locals() and nyuryoku_param_header_cell else "None",
            "additional_condition_columns_filled": columns_filled,
            "border_only_columns_processed": columns_bordered,
            "groups_processed": rows_processed if 'rows_processed' in locals() else 0,
            "rows_deleted": rows_deleted if 'rows_deleted' in locals() else 0,
            "action_merges": action_merges if 'action_merges' in locals() else 0,
            "koumoku_merges": koumoku_merges if 'koumoku_merges' in locals() else 0,
            "sousa_merges": sousa_merges if 'sousa_merges' in locals() else 0,
            "status": "success"
        }
        
        if progress_callback:
            progress_callback("Hoàn thành!")
        
        return result
        
    except Exception as e:
        if progress_callback:
            progress_callback(f"Lỗi: {str(e)}")
        raise e


def validate_file_compatibility(test_file_path, doc_file_path):
    """
    Check if the uploaded files are compatible for data copying
    
    Args:
        test_file_path (str): Path to test file
        doc_file_path (str): Path to documentation file
        
    Returns:
        dict: Validation result with status and message
    """
    try:
        test_ext = os.path.splitext(test_file_path)[1].lower()
        doc_ext = os.path.splitext(doc_file_path)[1].lower()
        
        # Define supported combinations
        supported_combinations = [
            # Excel to Excel combinations
            ('.xlsx', '.xlsx'), ('.xlsx', '.xls'),
            ('.xls', '.xlsx'), ('.xls', '.xls'),
            # Test file types that can work with Excel docs
            ('.java', '.xlsx'), ('.java', '.xls'),
            ('.py', '.xlsx'), ('.py', '.xls'),
            ('.js', '.xlsx'), ('.js', '.xls'),
            ('.ts', '.xlsx'), ('.ts', '.xls'),
            # Test file types that can work with Word docs  
            ('.java', '.docx'), ('.java', '.doc'),
            ('.py', '.docx'), ('.py', '.doc'),
            # Excel can work with Word docs
            ('.xlsx', '.docx'), ('.xlsx', '.doc'),
            ('.xls', '.docx'), ('.xls', '.doc'),
            # Word can work with Excel docs
            ('.docx', '.xlsx'), ('.docx', '.xls'),
            ('.doc', '.xlsx'), ('.doc', '.xls'),
            # Text files can work with any documentation
            ('.txt', '.xlsx'), ('.txt', '.xls'),
            ('.txt', '.docx'), ('.txt', '.doc'),
            ('.txt', '.txt')
        ]
        
        if (test_ext, doc_ext) in supported_combinations:
            return {
                "compatible": True,
                "message": f"File {test_ext} và {doc_ext} tương thích!"
            }
        else:
            return {
                "compatible": False,
                "message": f"File {test_ext} và {doc_ext} chưa được hỗ trợ!"
            }
            
    except Exception as e:
        return {
            "compatible": False,
            "message": f"Lỗi kiểm tra tương thích: {str(e)}"
        }


def get_copy_preview(test_file_path, doc_file_path):
    """
    Generate a preview of what will be copied
    
    Args:
        test_file_path (str): Path to mock documentation file  
        doc_file_path (str): Path to unit test file
        
    Returns:
        str: Preview text of the copy operation
    """
    try:
        test_name = os.path.basename(test_file_path)
        doc_name = os.path.basename(doc_file_path)
        
        # Try to get information about the source sheet
        preview_info = ""
        try:
            source_wb = load_workbook(test_file_path, data_only=True)
            sheet_name = "項目一覧"
            
            if sheet_name in source_wb.sheetnames:
                source_ws = source_wb[sheet_name]
                max_row = source_ws.max_row
                max_col = source_ws.max_column
                start_source_row = 5
                rows_to_copy = max_row - start_source_row + 1 if max_row >= start_source_row else 0
                preview_info = f"""
 Thông tin sheet nguồn:
• Sheet: {sheet_name}
• Tổng số dòng trong sheet: {max_row}
• Copy từ dòng: {start_source_row} đến {max_row}
• Số dòng sẽ copy: {rows_to_copy}
• Số cột: {max_col}
• Dữ liệu sẽ được paste từ dòng 6 trong file đích
                """
                source_wb.close()
            else:
                preview_info = f"\n Không tìm thấy sheet '{sheet_name}' trong file mocks"
                
        except Exception as e:
            preview_info = f"\n Không thể đọc thông tin sheet: {str(e)}"
        
        preview = f"""
📋 Thông tin sao chép dữ liệu:

 File mocks (nguồn): {test_name}
 File unit test (đích): {doc_name}

 Các thao tác sẽ thực hiện:
• Đọc dữ liệu từ dòng 5 trở đi trong sheet "項目一覧" của file mocks
• Copy cả giá trị và formatting (borders, colors, fonts, merged cells)
• Paste vào sheet "項目一覧" của file unit test với đầy đủ style
• Vị trí paste: Bắt đầu từ dòng thứ 6
• Thêm outside border bao quanh vùng dữ liệu đã copy
• Copy toàn bộ dữ liệu từ sheet "data_1" và paste vào cột tiếp theo (dòng 5)
• Giữ nguyên cấu trúc merged cells (cells đã gộp)
• Tạo bảng với thin border bên dưới data_1 (từ dòng 9)
• Thêm thin border bên trong bảng để tạo lưới
• Xóa sheet "data_1" sau khi copy xong để dọn dẹp{preview_info}

 Trạng thái: Sẵn sàng để thực hiện
        """
        
        return preview.strip()
        
    except Exception as e:
        return f"Lỗi tạo preview: {str(e)}" 


def get_action_copy_preview(test_file_path, doc_file_path):
    """
    Generate a preview of what will be copied for action data
    
    Args:
        test_file_path (str): Path to mock documentation file  
        doc_file_path (str): Path to unit test file
        
    Returns:
        str: Preview text of the copy operation
    """
    try:
        test_name = os.path.basename(test_file_path)
        doc_name = os.path.basename(doc_file_path)
        
        # Try to get information about the source sheet
        preview_info = ""
        try:
            source_wb = load_workbook(test_file_path, data_only=True)
            sheet_name = "アクション一覧"
            
            if sheet_name in source_wb.sheetnames:
                source_ws = source_wb[sheet_name]
                
                # Find the "アクションNo." header cell
                header_found = False
                header_row = None
                header_col = None
                data_count = 0
                
                for row in range(1, source_ws.max_row + 1):
                    for col in range(1, source_ws.max_column + 1):
                        cell = source_ws.cell(row=row, column=col)
                        if cell.value and str(cell.value).strip() == "アクションNo.":
                            header_found = True
                            header_row = row
                            header_col = col
                            
                            # Count data below header
                            for data_row in range(header_row + 1, source_ws.max_row + 1):
                                data_cell = source_ws.cell(row=data_row, column=header_col)
                                if data_cell.value is not None:
                                    data_count += 1
                                else:
                                    break
                            break
                    if header_found:
                        break
                
                if header_found:
                    preview_info = f"""
 Thông tin sheet nguồn:
• Sheet: {sheet_name}
• Header "アクションNo." tìm thấy tại: Dòng {header_row}, Cột {header_col}
• Số dữ liệu sẽ copy: {data_count} items
• Dữ liệu sẽ được paste vào cột "アクションNo." của file đích
                    """
                else:
                    preview_info = f"\n Không tìm thấy cột 'アクションNo.' trong sheet {sheet_name}"
                
                source_wb.close()
            else:
                preview_info = f"\n Không tìm thấy sheet '{sheet_name}' trong file mocks"
                
        except Exception as e:
            preview_info = f"\n Không thể đọc thông tin sheet: {str(e)}"
        
        preview = f"""
📋 Thông tin sao chép dữ liệu Action:

 File mocks (nguồn): {test_name}
 File unit test (đích): {doc_name}

 Các thao tác sẽ thực hiện:
• Tìm header "アクションNo." trong sheet "アクション一覧" của file mocks
• Copy toàn bộ dữ liệu bên dưới header này (cùng cột)
• Paste dữ liệu vào bên dưới header trong file unit test
• Copy dữ liệu sang cột "アクションNo." (không merge consecutive cells)
• Tìm cột "アクション" và copy dữ liệu sang 2 cột "項目" và "操作" (không merge consecutive cells)
• Tìm cột "処理条件" và copy dữ liệu sang cột "処理条件①" (không merge)
• Tìm cột "API URL" và copy dữ liệu sang cột "WEBAPI" (không merge)
• Tìm cột chứa "処理No." và copy dữ liệu sang cột "No." (không merge, chỉ giữ ký tự cuối)
• Kết hợp dữ liệu từ "処理概要" và "入力パラメータ" sang "想定結果" (chỉ kết hợp khi "入力パラメータ" khác rỗng và khác "-")
• Điền giá trị "-" và border vào các cột có sẵn "処理条件②" đến "処理条件⑥"
• Thêm border vào các cột có sẵn "実施者", "実施日", "結果" (không điền dữ liệu)
• Copy dữ liệu giữ nguyên như file mocks (không merge consecutive cells)
• Gộp hàng có cùng "No.", "アクションNo." và dòng đầu "想定結果", kết hợp dòng thứ 2 vào hàng đầu tiên, xóa hàng trùng
• Merge cells trong cột "アクションNo." cho các hàng có cùng ActionNo (chỉ merge cột này)
• Merge cells trong cột "項目" cho các hàng có cùng 項目 (chỉ merge cột này)
• Merge cells trong cột "操作" cho các hàng có cùng 操作 (chỉ merge cột này)
• Copy cả formatting (colors, fonts, borders)
• Nếu không tìm thấy header trong file đích, sẽ tạo mới tại A1{preview_info}

 Trạng thái: Sẵn sàng để thực hiện
        """
        
        return preview.strip() 
        
    except Exception as e:
        return f"Lỗi tạo preview: {str(e)}" 