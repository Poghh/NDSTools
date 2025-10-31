import os
import glob
from pathlib import Path
import tkinter as tk
from tkinter import messagebox

# Thử import xlwings để tương tác trực tiếp với Excel
try:
    import xlwings as xw
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    # Fallback to openpyxl
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.views import SheetView, Selection


def find_screen_folders(screen_list, base_folder):
    """
    Tìm các folder tương ứng với danh sách màn hình
    
    Args:
        screen_list (list): Danh sách tên màn hình
        base_folder (str): Đường dẫn folder tổng
        
    Returns:
        dict: Dictionary mapping screen name to folder path
    """
    found_folders = {}
    
    if not os.path.exists(base_folder):
        return found_folders
    
    # Lấy tất cả folder con
    for item in os.listdir(base_folder):
        item_path = os.path.join(base_folder, item)
        if os.path.isdir(item_path):
            # Kiểm tra xem folder có match với screen nào không
            for screen in screen_list:
                screen_clean = screen.strip()
                if screen_clean and (screen_clean.lower() in item.lower() or item.lower() in screen_clean.lower()):
                    found_folders[screen_clean] = item_path
                    break
    
    return found_folders


def format_excel_file_xlwings(file_path, output_callback=None):
    """
    Format Excel file using xlwings (preserves linked objects)
    """
    try:
        if output_callback:
            output_callback(f"📝 Đang xử lý với xlwings: {os.path.basename(file_path)}")
        
        # Mở Excel application (ẩn)
        app = xw.App(visible=False, add_book=False)
        
        try:
            # Mở workbook
            wb = app.books.open(file_path)
            
            # Format tất cả worksheet
            for i, sheet in enumerate(wb.sheets):
                try:
                    # Activate sheet
                    sheet.activate()
                    
                    # 1. Chọn cell A1
                    sheet.range('A1').select()
                    
                    # 2. Đặt zoom 100% - sử dụng COM API trực tiếp
                    try:
                        # Lấy window hiện tại
                        xl_app = app.api
                        active_window = xl_app.ActiveWindow
                        
                        # Set zoom
                        active_window.Zoom = 100
                        
                        # Set scroll position
                        active_window.ScrollRow = 1
                        active_window.ScrollColumn = 1
                        
                    except Exception as com_error:
                        if output_callback:
                            output_callback(f"    ⚠️ COM API lỗi: {str(com_error)}")
                    
                    if output_callback:
                        output_callback(f"  ✅ Sheet '{sheet.name}': A1, Zoom 100%, Scroll Top-Left")
                        
                except Exception as sheet_error:
                    if output_callback:
                        output_callback(f"  ⚠️ Sheet '{sheet.name}': Lỗi {str(sheet_error)}")
            
            # Đặt sheet đầu tiên làm active cuối cùng
            if wb.sheets:
                wb.sheets[0].activate()
                wb.sheets[0].range('A1').select()
                if output_callback:
                    output_callback(f"  🎯 Đặt sheet '{wb.sheets[0].name}' làm active")
            
            # Lưu và đóng
            wb.save()
            wb.close()
            
        finally:
            # Đóng Excel application
            app.quit()
        
        if output_callback:
            output_callback(f"✅ Hoàn thành (xlwings): {os.path.basename(file_path)}\n")
        
        return True
        
    except Exception as e:
        if output_callback:
            output_callback(f"❌ Lỗi xlwings {os.path.basename(file_path)}: {str(e)}\n")
        return False


def format_excel_file(file_path, output_callback=None, force_format=False):
    """
    Format Excel file - tự động chọn phương thức tốt nhất
    """
    if XLWINGS_AVAILABLE:
        # Ưu tiên xlwings vì bảo vệ linked objects tốt hơn
        return format_excel_file_xlwings(file_path, output_callback)
    else:
        # Fallback to openpyxl
        return format_excel_file_openpyxl(file_path, output_callback, force_format)


def format_excel_file_openpyxl(file_path, output_callback=None, force_format=False):
    """
    Format Excel file using openpyxl (fallback method)
    
    Args:
        file_path (str): Đường dẫn file Excel
        output_callback (function): Callback để hiển thị log
        force_format (bool): Có ép buộc format ngay cả khi có objects không
        
    Returns:
        bool: True nếu thành công, False nếu có lỗi
    """
    try:
        if output_callback:
            output_callback(f"📝 Đang xử lý: {os.path.basename(file_path)}")
        
        # Mở workbook để kiểm tra
        workbook = openpyxl.load_workbook(file_path)
        
        # Kiểm tra xem file có chứa objects/images không
        has_objects = False
        object_count = 0
        
        for sheet in workbook.worksheets:
            if hasattr(sheet, '_images') and sheet._images:
                has_objects = True
                object_count += len(sheet._images)
            if hasattr(sheet, '_charts') and sheet._charts:
                has_objects = True
                object_count += len(sheet._charts)
        
        if has_objects and not force_format:
            if output_callback:
                output_callback(f"  ⚠️ PHÁT HIỆN {object_count} objects/images trong file")
                output_callback(f"  🛡️ Để bảo vệ linked objects, tool sẽ BỎ QUA file này")
                output_callback(f"  💡 Bật 'Force Format' nếu muốn xử lý dù có rủi ro")
            
            workbook.close()
            return True  # Bỏ qua nhưng không báo lỗi
        
        elif has_objects and force_format:
            if output_callback:
                output_callback(f"  ⚠️ PHÁT HIỆN {object_count} objects/images - NHƯNG VẪN XỬ LÝ")
                output_callback(f"  🚨 CẢNH BÁO: Linked objects có thể bị mất!")
        
        # Tiến hành format (có hoặc không có objects tùy theo force_format)
        try:
            # 1. Format tất cả sheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                try:
                    # Tạo sheet_view nếu cần
                    if not hasattr(worksheet, 'sheet_view') or worksheet.sheet_view is None:
                        worksheet.sheet_view = SheetView()
                    
                    # Format view properties
                    worksheet.sheet_view.zoomScale = 100
                    selection = Selection(activeCell='A1', sqref='A1')
                    worksheet.sheet_view.selection = [selection]
                    worksheet.sheet_view.topLeftCell = 'A1'
                    
                    if output_callback:
                        output_callback(f"  ✅ Sheet '{sheet_name}': A1, Zoom 100%, Scroll Top-Left")
                        
                except Exception as sheet_error:
                    if output_callback:
                        output_callback(f"  ⚠️ Sheet '{sheet_name}': Lỗi {str(sheet_error)}")
            
            # 2. Đặt sheet đầu tiên làm active
            if workbook.sheetnames:
                first_sheet_name = workbook.sheetnames[0]
                workbook.active = workbook[first_sheet_name]
                
                if output_callback:
                    output_callback(f"  🎯 Đặt sheet '{first_sheet_name}' làm active")
            
            # 3. Lưu file
            workbook.save(file_path)
            workbook.close()
            
        except Exception as e:
            workbook.close()
            raise e
        
        if output_callback:
            output_callback(f"✅ Hoàn thành: {os.path.basename(file_path)}\n")
        
        return True
        
    except Exception as e:
        if output_callback:
            output_callback(f"❌ Lỗi xử lý {os.path.basename(file_path)}: {str(e)}\n")
        return False


def process_screen_folders(screen_list, base_folder, output_callback=None, force_format=False):
    """
    Xử lý format Excel cho danh sách màn hình
    
    Args:
        screen_list (list): Danh sách tên màn hình
        base_folder (str): Đường dẫn folder tổng
        output_callback (function): Callback để hiển thị log
        force_format (bool): Có ép buộc format ngay cả khi có objects không
        
    Returns:
        dict: Kết quả xử lý {screen: {success_count, error_count, files}}
    """
    results = {}
    
    if output_callback:
        output_callback("🔍 Bắt đầu tìm kiếm folder màn hình...\n")
    
    # Tìm folder tương ứng
    found_folders = find_screen_folders(screen_list, base_folder)
    
    if not found_folders:
        if output_callback:
            output_callback("❌ Không tìm thấy folder nào phù hợp!\n")
        return results
    
    if output_callback:
        output_callback(f"✅ Tìm thấy {len(found_folders)} folder:\n")
        for screen, folder_path in found_folders.items():
            output_callback(f"  📁 {screen} → {folder_path}\n")
        output_callback("\n")
    
    # Xử lý từng folder
    for screen_name, folder_path in found_folders.items():
        if output_callback:
            output_callback(f"📂 Xử lý folder: {screen_name}\n")
        
        # Tìm tất cả file Excel trong folder (bao gồm cả subfolder)
        excel_files = []
        for ext in ['*.xlsx', '*.xls']:
            # Sử dụng recursive=True để tìm trong tất cả subfolder
            found_files = glob.glob(os.path.join(folder_path, '**', ext), recursive=True)
            excel_files.extend(found_files)
        
        # Loại bỏ file trùng lặp bằng cách chuyển thành set rồi về list
        excel_files = list(set(excel_files))
        
        success_count = 0
        error_count = 0
        processed_files = []
        
        if not excel_files:
            if output_callback:
                output_callback(f"  ⚠️ Không tìm thấy file Excel nào trong {screen_name}\n")
        else:
            if output_callback:
                output_callback(f"  📊 Tìm thấy {len(excel_files)} file Excel\n")
                # Debug: hiển thị danh sách file tìm được
                for i, file_path in enumerate(excel_files, 1):
                    file_name = os.path.basename(file_path)
                    output_callback(f"    {i}. {file_name}\n")
            
            for excel_file in excel_files:
                # Bỏ qua file tạm của Excel (bắt đầu với ~$)
                if os.path.basename(excel_file).startswith('~$'):
                    continue
                
                if format_excel_file(excel_file, output_callback, force_format):
                    success_count += 1
                else:
                    error_count += 1
                
                processed_files.append(excel_file)
        
        results[screen_name] = {
            'success_count': success_count,
            'error_count': error_count,
            'files': processed_files,
            'folder_path': folder_path
        }
        
        if output_callback:
            output_callback(f"📊 Kết quả {screen_name}: {success_count} thành công, {error_count} lỗi\n")
            output_callback("-" * 50 + "\n")
    
    return results


def validate_inputs(screen_text, base_folder):
    """
    Validate input từ user
    
    Args:
        screen_text (str): Text chứa danh sách màn hình
        base_folder (str): Đường dẫn folder tổng
        
    Returns:
        tuple: (is_valid, error_message, screen_list)
    """
    # Kiểm tra folder tổng
    if not base_folder or not base_folder.strip():
        return False, "Vui lòng chọn folder tổng!", []
    
    if not os.path.exists(base_folder):
        return False, f"Folder không tồn tại: {base_folder}", []
    
    if not os.path.isdir(base_folder):
        return False, f"Đường dẫn không phải là folder: {base_folder}", []
    
    # Kiểm tra danh sách màn hình
    if not screen_text or not screen_text.strip():
        return False, "Vui lòng nhập danh sách màn hình!", []
    
    # Parse danh sách màn hình (tách theo dòng hoặc dấu phẩy)
    screen_list = []
    lines = screen_text.strip().split('\n')
    for line in lines:
        # Tách theo dấu phẩy nếu có
        if ',' in line:
            screens = [s.strip() for s in line.split(',') if s.strip()]
            screen_list.extend(screens)
        else:
            screen = line.strip()
            if screen:
                screen_list.append(screen)
    
    if not screen_list:
        return False, "Danh sách màn hình trống!", []
    
    return True, "", screen_list
