"""
Module xử lý dữ liệu CareBase từ file Excel
Logic xử lý sẽ được thêm sau
"""

import pandas as pd
import re


def create_empty_carebase_template(output_callback=None):
    """
    Tạo DataFrame trống với các cột header cho CareBase
    
    Các cột header:
    項番, 取消, 優先度, STEP, (Folder No.), Id màn hình, Tên màn hình, 
    Tên ticket, Ticket cha, Nội dung đối ứng, Trạng thái bug, Trạng thái ticket,
    Cứ điểm tạo, Người tạo, Ngày phát sinh, Người đối ứng, Ngày done mong muốn,
    Người trả lời, Ngày trả lời, Người check, Ngày check, Phân loại bug, Ghi chú
    
    Args:
        output_callback (callable): Hàm callback để hiển thị log (optional)
    
    Returns:
        pd.DataFrame: DataFrame trống với các cột header
    """
    if output_callback:
        output_callback(" 🔄 Đang tạo template với header...\n")
    
    # Định nghĩa các cột header
    columns = [
        "項番",
        "取消",
        "優先度",
        "STEP",
        "(Folder No.)",
        "Id màn hình",
        "Tên màn hình",
        "Tên ticket",
        "Ticket cha",
        "Nội dung đối ứng",
        "Trạng thái bug",
        "Trạng thái ticket",
        "Cứ điểm tạo",
        "Người tạo",
        "Ngày phát sinh",
        "Người đối ứng",
        "Ngày done mong muốn",
        "Người trả lời",
        "Ngày trả lời",
        "Người check",
        "Ngày check",
        "Phân loại bug",
        "Ghi chú",
    ]
    
    # Tạo DataFrame trống với các cột header
    empty_df = pd.DataFrame(columns=columns)
    
    if output_callback:
        output_callback(f" ✅ Đã tạo template với {len(columns)} cột\n")
        output_callback(" 📝 Danh sách cột:\n")
        for idx, col in enumerate(columns, 1):
            output_callback(f"   {idx}. {col}\n")
        output_callback("\n")
    
    return empty_df


def extract_third_bracket_value(text):
    """
    Trích xuất giá trị trong ngoặc vuông thứ 3 từ text
    
    Ví dụ: 【カイマデータ】【結合テスト】【1】【遷移図_1】【GUI02259_適用事業所の選択】返却データに関する不具合
    Kết quả: 1
    
    Args:
        text: Chuỗi text cần xử lý
    
    Returns:
        str: Giá trị trong ngoặc vuông thứ 3, hoặc rỗng nếu không tìm thấy
    """
    if pd.isna(text) or not isinstance(text, str):
        return ""
    
    # Tìm tất cả các giá trị trong ngoặc vuông 【】
    pattern = r'【([^】]+)】'
    matches = re.findall(pattern, text)
    
    # Lấy giá trị thứ 3 (index 2)
    if len(matches) >= 3:
        return matches[2].strip()
    
    return ""


def extract_fourth_bracket_after_underscore(text):
    """
    Trích xuất giá trị sau dấu "_" trong ngoặc vuông thứ 4 từ text
    
    Ví dụ: 【カイマデータ】【結合テスト】【1】【遷移図_1】【GUI02259_適用事業所の選択】返却データに関する不具合
    Ngoặc thứ 4: 【遷移図_1】
    Kết quả: 1
    
    Args:
        text: Chuỗi text cần xử lý
    
    Returns:
        str: Giá trị sau dấu "_" trong ngoặc vuông thứ 4, hoặc rỗng nếu không tìm thấy
    """
    if pd.isna(text) or not isinstance(text, str):
        return ""
    
    # Tìm tất cả các giá trị trong ngoặc vuông 【】
    pattern = r'【([^】]+)】'
    matches = re.findall(pattern, text)
    
    # Lấy giá trị thứ 4 (index 3)
    if len(matches) >= 4:
        fourth_value = matches[3].strip()
        # Tìm giá trị sau dấu "_"
        if "_" in fourth_value:
            parts = fourth_value.split("_", 1)
            if len(parts) > 1:
                return parts[1].strip()
    
    return ""


def extract_fifth_bracket_split_by_underscore(text):
    """
    Trích xuất và tách giá trị trong ngoặc vuông thứ 5 từ text theo dấu "_"
    
    Ví dụ: 【カイマデータ】【結合テスト】【1】【遷移図_1】【GUI02259_適用事業所の選択】返却データに関する不具合
    Ngoặc thứ 5: 【GUI02259_適用事業所の選択】
    Kết quả: ("GUI02259", "適用事業所の選択")
    
    Args:
        text: Chuỗi text cần xử lý
    
    Returns:
        tuple: (phần trước "_", phần sau "_"), hoặc ("", "") nếu không tìm thấy
    """
    if pd.isna(text) or not isinstance(text, str):
        return ("", "")
    
    # Tìm tất cả các giá trị trong ngoặc vuông 【】
    pattern = r'【([^】]+)】'
    matches = re.findall(pattern, text)
    
    # Lấy giá trị thứ 5 (index 4)
    if len(matches) >= 5:
        fifth_value = matches[4].strip()
        # Tách theo dấu "_"
        if "_" in fifth_value:
            parts = fifth_value.split("_", 1)
            if len(parts) == 2:
                return (parts[0].strip(), parts[1].strip())
            elif len(parts) == 1:
                return (parts[0].strip(), "")
    
    return ("", "")


def extract_text_after_fifth_bracket(text):
    """
    Trích xuất text sau ngoặc vuông thứ 5 từ text
    
    Ví dụ: 【カイマデータ】【結合テスト】【1】【遷移図_1】【GUI02259_適用事業所の選択】返却データに関する不具合
    Kết quả: 返却データに関する不具合
    
    Args:
        text: Chuỗi text cần xử lý
    
    Returns:
        str: Text sau ngoặc vuông thứ 5, hoặc rỗng nếu không tìm thấy
    """
    if pd.isna(text) or not isinstance(text, str):
        return ""
    
    # Tìm tất cả các ngoặc vuông 【】
    pattern = r'【[^】]+】'
    matches = list(re.finditer(pattern, text))
    
    # Nếu có ít nhất 5 ngoặc vuông
    if len(matches) >= 5:
        # Lấy vị trí kết thúc của ngoặc vuông thứ 5
        fifth_bracket_end = matches[4].end()
        # Lấy text sau ngoặc vuông thứ 5
        text_after = text[fifth_bracket_end:].strip()
        return text_after
    
    return ""


def format_author_name(author_name):
    """
    Format tên tác giả theo quy tắc:
    - "NT Duy" -> "KMD DuyNT"
    - Các trường hợp khác: KMD + tên cuối (chữ đầu viết hoa) + các chữ đầu của các từ còn lại (viết hoa)
      Ví dụ: "doan van huynh" -> "KMD HuynhDV"
    
    Args:
        author_name: Tên tác giả cần format
    
    Returns:
        str: Tên đã được format
    """
    if pd.isna(author_name) or not isinstance(author_name, str):
        return ""
    
    author_name = author_name.strip()
    
    # Xử lý trường hợp đặc biệt
    if author_name == "NT Duy":
        return "KMD DuyNT"
    elif author_name.lower() == "thao tran thi":
        return "KMD ThaoTT"
    
    # Xử lý các trường hợp khác
    # Tách thành các từ
    words = author_name.split()
    
    if len(words) == 0:
        return ""
    
    # Lấy tên cuối cùng (từ cuối)
    last_name = words[-1]
    # Viết hoa chữ đầu của tên cuối
    last_name_formatted = last_name.capitalize()
    
    # Lấy các chữ đầu của các từ còn lại (bỏ qua tên cuối)
    first_letters = ""
    for word in words[:-1]:
        if word:
            first_letters += word[0].upper()
    
    # Ghép lại: KMD + tên cuối + các chữ đầu
    result = f"KMD {last_name_formatted}{first_letters}"
    
    return result


def process_carebase_data(df: pd.DataFrame, output_callback=None):
    """
    Xử lý dữ liệu CareBase từ DataFrame
    
    Rule:
    - Bắt đầu từ dòng 2 của input (index 1)
    - Lấy data từ cột "Subject"
    - Trích xuất giá trị trong ngoặc vuông thứ 3 【】 -> điền vào cột "STEP"
    - Trích xuất giá trị sau "_" trong ngoặc vuông thứ 4 【】 -> điền vào cột "(Folder No.)"
    - Trích xuất và tách giá trị trong ngoặc vuông thứ 5 【】 theo "_" -> 
      phần trước điền vào "Id màn hình", phần sau điền vào "Tên màn hình"
    - Trích xuất text sau ngoặc vuông thứ 5 -> điền vào cột "Nội dung đối ứng"
    
    Args:
        df (pd.DataFrame): DataFrame chứa dữ liệu từ file Excel
        output_callback (callable): Hàm callback để hiển thị log (optional)
    
    Returns:
        pd.DataFrame: DataFrame đã được xử lý với dữ liệu từ input
    """
    if output_callback:
        output_callback(" 🔄 Đang xử lý dữ liệu từ input...\n")
    
    # Tạo DataFrame trống với header
    processed_df = create_empty_carebase_template(output_callback)
    
    # Kiểm tra xem có cột "Subject" không
    if "Subject" not in df.columns:
        if output_callback:
            output_callback(" ⚠️  Không tìm thấy cột 'Subject' trong file input\n")
            output_callback(" 📋 Các cột có sẵn:\n")
            for col in df.columns:
                output_callback(f"   - {col}\n")
        return processed_df
    
    # Xử lý từng dòng từ dòng 2 (index 1) trở đi
    output_rows = []
    
    if output_callback:
        output_callback(f" 📊 Bắt đầu xử lý {len(df)} dòng từ input...\n")
        output_callback(f" 📋 Số dòng dữ liệu (sau header): {len(df)}\n")
    
    # Lấy tất cả các dòng dữ liệu (bỏ qua header nếu có)
    # pandas đọc Excel/CSV thường có header ở dòng đầu, data bắt đầu từ index 0
    # Nhưng nếu header đã được đọc, thì data bắt đầu từ index 0
    row_count = 0
    for idx, row in df.iterrows():
        row_count += 1
        
        # Lấy giá trị từ cột "Subject"
        subject_value = row.get("Subject", "")
        
        # Lấy giá trị từ cột "#" -> Tên ticket
        ticket_name = row.get("#", "")
        if pd.notna(ticket_name):
            ticket_name = str(ticket_name).strip()
        else:
            ticket_name = ""
        
        # Lấy giá trị từ cột "Parent task" -> Ticket cha
        parent_task = row.get("Parent task", "")
        if pd.notna(parent_task):
            # Chuyển đổi số float thành int nếu có thể, để tránh .0
            try:
                # Thử chuyển thành float rồi int
                if isinstance(parent_task, (int, float)):
                    parent_task = str(int(float(parent_task)))
                else:
                    parent_task = str(parent_task).strip()
                    # Nếu là số dạng string có .0, loại bỏ
                    if parent_task.endswith('.0'):
                        parent_task = parent_task[:-2]
            except (ValueError, TypeError):
                parent_task = str(parent_task).strip()
        else:
            parent_task = ""
        
        # Lấy giá trị từ cột "Status" -> Trạng thái bug
        status_value = row.get("Status", "")
        if pd.notna(status_value):
            status_value = str(status_value).strip()
        else:
            status_value = ""
        
        # Lấy giá trị từ cột "Author" -> Người tạo
        author_value = row.get("Author", "")
        if pd.notna(author_value):
            author_value = format_author_name(str(author_value))
        else:
            author_value = ""
        
        # Trích xuất giá trị trong ngoặc vuông thứ 3 -> STEP
        step_value = extract_third_bracket_value(subject_value)
        
        # Trích xuất giá trị sau "_" trong ngoặc vuông thứ 4 -> (Folder No.)
        folder_no_value = extract_fourth_bracket_after_underscore(subject_value)
        
        # Trích xuất và tách giá trị trong ngoặc vuông thứ 5 -> Id màn hình và Tên màn hình
        screen_id, screen_name = extract_fifth_bracket_split_by_underscore(subject_value)
        
        # Trích xuất text sau ngoặc vuông thứ 5 -> Nội dung đối ứng
        content_value = extract_text_after_fifth_bracket(subject_value)
        
        # Tạo dòng mới cho output
        output_row = {col: "" for col in processed_df.columns}
        output_row["STEP"] = step_value
        output_row["(Folder No.)"] = folder_no_value
        output_row["Id màn hình"] = screen_id
        output_row["Tên màn hình"] = screen_name
        output_row["Tên ticket"] = ticket_name
        output_row["Ticket cha"] = parent_task
        output_row["Nội dung đối ứng"] = content_value
        output_row["Trạng thái bug"] = status_value
        
        # Xác định Trạng thái ticket dựa trên Trạng thái bug
        if status_value == "終了":
            ticket_status = "終了"
        else:
            ticket_status = "対応中"
        output_row["Trạng thái ticket"] = ticket_status
        
        # Cứ điểm tạo luôn là "ベトナム"
        output_row["Cứ điểm tạo"] = "ベトナム"
        
        # Người tạo từ cột Author
        output_row["Người tạo"] = author_value
        
        output_rows.append(output_row)
        
        if output_callback and row_count % 100 == 0:
            output_callback(f"   Đã xử lý {row_count} dòng...\n")
    
    # Thêm các dòng vào DataFrame
    if output_rows:
        new_df = pd.DataFrame(output_rows)
        processed_df = pd.concat([processed_df, new_df], ignore_index=True)
        
        if output_callback:
            output_callback(f" ✅ Đã xử lý {len(output_rows)} dòng từ input\n")
            output_callback(f" 📊 Tổng số dòng output: {len(processed_df)}\n")
    else:
        if output_callback:
            output_callback(" ⚠️  Không có dữ liệu để xử lý\n")
    
    return processed_df


def save_processed_data(df: pd.DataFrame, output_path: str, output_callback=None):
    """
    Lưu dữ liệu đã xử lý ra file Excel với header có background màu xanh dương
    
    Args:
        df (pd.DataFrame): DataFrame đã được xử lý
        output_path (str): Đường dẫn file output
        output_callback (callable): Hàm callback để hiển thị log (optional)
    
    Returns:
        str: Đường dẫn file đã lưu
    """
    if output_callback:
        output_callback(f" 💾 Đang lưu kết quả vào: {output_path}\n")
    
    # Lưu DataFrame vào Excel
    df.to_excel(output_path, index=False, engine='openpyxl')
    
    # Format header với background màu xanh dương
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Màu xanh dương cho background
        blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        black_font = Font(color="000000", bold=True)  # Chữ đen, đậm
        
        # Format hàng header (hàng 1)
        for cell in ws[1]:
            cell.fill = blue_fill
            cell.font = black_font
        
        wb.save(output_path)
        
        if output_callback:
            output_callback(f" ✅ Đã lưu thành công với header màu xanh dương!\n")
    except Exception as e:
        if output_callback:
            output_callback(f" ⚠️  Đã lưu file nhưng không thể format header: {str(e)}\n")
        else:
            # Nếu không có callback, vẫn lưu được file
            pass
    
    return output_path

