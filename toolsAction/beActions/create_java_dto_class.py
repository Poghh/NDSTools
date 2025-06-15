import tkinter as tk

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


def generated_dto(tab_instance):
    if not tab_instance.file_path:
        print("Bạn chưa chọn file Excel!")
        return
    tab_instance.output_text.delete(1.0, tk.END)
    request_sheet, response_sheet = find_request_and_response_sheets(tab_instance.file_path)

    for sheet in [request_sheet, response_sheet]:
        if sheet is None:
            continue

        header_ranges, header_row = find_header_ranges_in_sheet(sheet)
        data_by_header, hierarchy = extract_column_data_by_headers(sheet, header_ranges, header_row)

        headers = {
            "フィールド名": data_by_header["フィールド名"],
            "データ構造": data_by_header["データ構造"],
            "データタイプ": data_by_header["データタイプ"],
            "必須": data_by_header["必須"],
        }

        java_code = convert_to_java_class(sheet.title, headers, hierarchy)

        tab_instance.output_text.insert(tk.END, f"\n=== Java class cho {sheet.title} ===\n\n")
        tab_instance.output_text.insert(tk.END, java_code)
        tab_instance.output_text.insert(tk.END, "\n\n" + "=" * 80 + "\n")


def find_request_and_response_sheets(file_path):
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        sheet_names = workbook.sheetnames

        request_sheet = next(
            (workbook[name] for name in sheet_names if "リクエストパラメータ詳細" in name),
            None,
        )
        response_sheet = next(
            (workbook[name] for name in sheet_names if "レスポンスパラメータ詳細" in name),
            None,
        )

        if not request_sheet:
            print("Không tìm thấy sheet リクエストパラメータ詳細")
        if not response_sheet:
            print("Không tìm thấy sheet レスポンスパラメータ詳細")

        return request_sheet, response_sheet

    except Exception as e:
        print(f"Lỗi khi mở file: {e}")
        return None, None


def find_header_ranges_in_sheet(sheet):
    target_headers = ["フィールド名", "データ構造", "必須", "データタイプ"]
    header_ranges = {}
    header_row = None

    for row in sheet.iter_rows(min_row=1, max_row=10):
        headers_in_row = [str(cell.value).strip() if cell.value else None for cell in row]
        if any(h in headers_in_row for h in target_headers):
            header_row = row[0].row
            current_header = None
            start_idx = None

            for idx, value in enumerate(headers_in_row):
                if value in target_headers:
                    if current_header is not None and start_idx is not None:
                        end_idx = idx - 1 if idx - 1 >= start_idx else idx
                        start_col = get_column_letter(start_idx + 1)
                        end_col = get_column_letter(end_idx + 1)
                        header_ranges[current_header] = (
                            f"{start_col}-{end_col}" if start_col != end_col else start_col
                        )
                    current_header = value
                    start_idx = idx
                elif value is None and current_header:
                    continue
                else:
                    if current_header and start_idx is not None:
                        end_idx = idx - 1 if idx - 1 >= start_idx else idx
                        start_col = get_column_letter(start_idx + 1)
                        end_col = get_column_letter(end_idx + 1)
                        header_ranges[current_header] = (
                            f"{start_col}-{end_col}" if start_col != end_col else start_col
                        )
                        current_header = None
                        start_idx = None

            if current_header and start_idx is not None:
                end_idx = len(headers_in_row) - 1
                start_col = get_column_letter(start_idx + 1)
                end_col = get_column_letter(end_idx + 1)
                header_ranges[current_header] = (
                    f"{start_col}-{end_col}" if start_col != end_col else start_col
                )
            break

    return header_ranges, header_row


def extract_column_data_by_headers(sheet, header_ranges, header_row):
    data_by_header = {}
    hierarchy = {}

    for header, col_range in header_ranges.items():
        start_col = column_index_from_string(col_range.split("-")[0])
        end_col = column_index_from_string(col_range.split("-")[-1])

        values = []
        for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
            if start_col == end_col:
                values.append(row[start_col - 1].value)
            else:
                merged_values = [row[col - 1].value for col in range(start_col, end_col + 1)]
                values.append(merged_values)

        data_by_header[header] = values

    if "データ構造" in header_ranges:
        start_col = column_index_from_string(header_ranges["データ構造"].split("-")[0])
        end_col = column_index_from_string(header_ranges["データ構造"].split("-")[-1])

        data_structure_cells = [
            [row[col - 1] for col in range(start_col, end_col + 1)]
            for row in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row)
        ]

        def detect_parent_child_structure(rows):
            hierarchy = {}
            parent_stack = []

            for row in rows:
                for cell in row:
                    if cell and cell.value is not None:
                        value = cell.value
                        indent = cell.alignment.indent if cell.alignment else 0

                        while parent_stack and indent <= parent_stack[-1][0]:
                            parent_stack.pop()

                        if parent_stack:
                            parent = parent_stack[-1][1]
                            hierarchy.setdefault(parent, []).append(value)

                        hierarchy.setdefault(value, [])
                        parent_stack.append((indent, value))
                        break
            return {k: v for k, v in hierarchy.items() if v}

        hierarchy = detect_parent_child_structure(data_structure_cells)

    return data_by_header, hierarchy


def convert_to_java_class(sheet_name: str, headers: dict, hierarchy: dict) -> str:
    class_name = "InDto" if "リクエスト" in sheet_name else "OutDto"
    result = []

    def field_comment(name: str):
        return f"/** {name} */"

    def camel_case(s: str):
        import re

        if not s:
            return s
        parts = re.split(r"[_\s]+", s)
        return parts[0].lower() + "".join(p.capitalize() for p in parts[1:])

    def pascal_case(s: str):
        return s[0].upper() + s[1:] if s else s

    result.append("@Getter\n@Setter")
    result.append(f"public class {class_name} extends IDtoImpl " + "{")
    result.append("    /** serialVersionUID. */")
    result.append("    private static final long serialVersionUID = 1L;\n")

    required_flags = headers.get("必須", [])
    data_types = headers.get("データタイプ", [])
    skip_field_names = {camel_case(parent) for parent in hierarchy}
    for children in hierarchy.values():
        skip_field_names.update(camel_case(child) for child in children)

    # === Flat fields ===
    for idx, (name, field) in enumerate(zip(headers["フィールド名"], headers["データ構造"])):
        field_var = field[0]
        if field_var in skip_field_names or field_var in [None, "None"]:
            continue

        required = idx < len(required_flags) and str(required_flags[idx]).strip() != "-"
        dtype = data_types[idx] if idx < len(data_types) else None
        if dtype in [
            "List<String>",
            "List<Integer>",
            "List<Boolean>",
            "List<Double>",
            "List<Float>",
        ]:
            java_type = "List<String>"
        else:
            java_type = "String"

        result.append(f"    {field_comment(name)}")
        if required:
            result.append("    @NotEmpty")
        result.append(f"    private {java_type} {field_var};\n")

    # === Hierarchical (nested class) fields ===
    for parent, children in hierarchy.items():
        parent_var = camel_case(parent)
        parent_class = pascal_case(parent)
        parent_comment = next(
            (
                name
                for name, field in zip(headers["フィールド名"], headers["データ構造"])
                if isinstance(field, list) and field[0] == parent
            ),
            parent,
        )
        parent_required = next(
            (
                required
                for field, required in zip(headers["データ構造"], headers.get("必須", []))
                if isinstance(field, list) and field[0] == parent
            ),
            "-",
        )
        parent_required_flag = str(parent_required).strip() != "-"
        parent_datatype_raw = next(
            (
                dtype
                for field, dtype in zip(headers["データ構造"], headers.get("データタイプ", []))
                if isinstance(field, list) and field[0] == parent
            ),
            None,
        )

        # Determine Java type
        if parent_datatype_raw in ["List<String>", "List<Integer>"]:
            java_type = "List<String>"
        elif parent_datatype_raw == "List":
            java_type = f"List<{parent_class}>"
        else:
            java_type = parent_class

        result.append(f"    {field_comment(parent_comment)}")
        if parent_required_flag:
            result.append("    @NotEmpty")
        result.append(f"    private {java_type} {parent_var};\n")

        result.append("    @Getter\n    @Setter")
        result.append(f"    public static class {parent_class} extends IDtoImpl " + "{")
        result.append("        /** serialVersionUID. */")
        result.append("        private static final long serialVersionUID = 1L;\n")
        for child in children:
            comment = ""
            field_name = camel_case(child)
            child_required = False
            for idx, (name, field) in enumerate(
                zip(headers["フィールド名"], headers["データ構造"])
            ):
                field_candidate = field[0] if isinstance(field, list) else field
                if field_candidate == child:
                    comment = name
                    field_name = camel_case(field_candidate)
                    if idx < len(required_flags) and str(required_flags[idx]).strip() != "-":
                        child_required = True
                    break

            result.append(f"        {field_comment(comment)}")
            if child_required:
                result.append("        @NotEmpty")
            result.append(f"        private String {field_name};\n")
        result.append("    }\n")

    result.append("}")
    return "\n".join(result)
