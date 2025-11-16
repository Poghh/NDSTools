import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from toolsAction.beActions.count_code import count_code
from toolsAction.beActions.process_selfcheck_excel import (
    process_selfcheck_excel,
)


class SelfCheckTab:
    def __init__(self, parent, style):
        self.parent = parent
        self.style = style

        self.frame = ttk.Frame(parent, style="Card.TFrame")
        self.frame.columnconfigure(0, weight=1)

        self._build_ui()

    def _build_ui(self):
        frame = self.frame

        ttk.Label(frame, text="Self Check", style="Header.TLabel").grid(
            row=0, column=0, sticky="w", pady=(2, 6), padx=10
        )
        frame.columnconfigure(0, weight=1)

        # --- Codes group ---
        codes_group = ttk.Labelframe(frame, text="Danh sách mã màn hình", style="Card.TLabelframe")
        codes_group.grid(row=1, column=0, sticky="ew", pady=(0, 8), padx=10)
        codes_group.columnconfigure(0, weight=1)

        ttk.Label(codes_group, text="Dán mã màn hình (GUIxxxxx)", style="Muted.TLabel").grid(
            row=0, column=0, sticky="w", pady=(4, 2), padx=8
        )

        self.codes_text = tk.Text(
            codes_group,
            height=7,
            relief="flat",
            highlightthickness=1,
            highlightbackground="#e5e7eb",
        )
        self.codes_text.grid(row=1, column=0, sticky="ew", padx=(8, 4), pady=(0, 8))

        # — Buttons —
        btns = tk.Frame(codes_group, bg=self.style.lookup("Card.TLabelframe", "background"))
        btns.grid(row=1, column=1, sticky="n", padx=(0, 8), pady=(0, 8))

        ttk.Button(btns, text="Dán\nclipboard", style="Secondary.TButton").pack(
            pady=(0, 4), fill="x"
        )

        ttk.Button(
            btns,
            text="Xoá\nhết",
            style="Secondary.TButton",
            command=lambda: self.codes_text.delete("1.0", tk.END),
        ).pack(fill="x")

        # --- Toolbar ---
        toolbar = ttk.Frame(frame, style="Card.TFrame")
        toolbar.grid(row=2, column=0, sticky="ew", pady=(4, 4), padx=10)

        # 4 cột chia đều nhau
        for i in range(4):
            toolbar.columnconfigure(i, weight=1)

        ttk.Button(
            toolbar,
            text="  📂 Chọn thư mục\n(chứa file Self Check)",
            style="Secondary.TButton",
        ).grid(row=0, column=0, padx=4, sticky="ew")

        ttk.Button(
            toolbar,
            text="📄 Tải File",
            style="Secondary.TButton",
        ).grid(row=0, column=1, padx=4, sticky="ew")

        ttk.Button(
            toolbar,
            text="📊 Đếm dòng code",
            style="Primary.TButton",
        ).grid(row=0, column=2, padx=4, sticky="ew")

        ttk.Button(
            toolbar,
            text="🧾 Xuất báo cáo",
            style="Primary.TButton",
        ).grid(row=0, column=3, padx=4, sticky="ew")

    def select_self_check_folder(self):
        """Chọn thư mục chứa các file Self Check, lọc theo mã GUI trong ô văn bản và xử lý từng file Excel."""
        # 1️⃣ Lấy danh sách mã GUIxxxxx từ vùng văn bản
        raw = self.screen_codes_text.get("1.0", tk.END)
        found = re.findall(
            r"(?<![A-Za-z0-9])(GUI\d{5}|[A-Z][A-Z0-9]{5})(?![A-Za-z0-9])", raw, flags=re.IGNORECASE
        )

        seen = set()
        self.screen_codes = []
        for c in (x.upper().strip() for x in found):
            if c and c not in seen:
                seen.add(c)
                self.screen_codes.append(c)

        if not self.screen_codes:
            messagebox.showinfo(
                "Chưa có danh sách mã",
                "Vui lòng dán danh sách mã (mỗi dòng 1 mã) vào ô phía trên trước khi chọn thư mục.",
            )
            return

        # 2️⃣ Chọn thư mục chứa file self-check
        folder_path = filedialog.askdirectory(title="Chọn thư mục Self Check")
        if not folder_path:
            return

        self.self_check_path = folder_path
        self.self_check_label.config(text=f"Thư mục: {os.path.basename(folder_path)}")

        # 3️⃣ Lấy danh sách file Excel trong thư mục
        valid_exts = (".xlsx", ".xls", ".xlsm", ".csv")
        all_files = [
            f
            for f in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, f)) and f.lower().endswith(valid_exts)
        ]
        if not all_files:
            messagebox.showwarning(
                "Không có file",
                "Không tìm thấy file Self Check (.xlsx/.xls/.xlsm/.csv) trong thư mục này.",
            )
            return

        # 4️⃣ Lọc file có chứa mã
        matched_files = []
        for f in all_files:
            fu = f.upper()
            if any(code in fu for code in self.screen_codes):
                matched_files.append(os.path.join(folder_path, f))

        if not matched_files:
            messagebox.showwarning(
                "Không tìm thấy", "Không có file nào trùng với mã màn hình đã dán."
            )
            return

        # 5️⃣ Lưu danh sách vào biến instance
        self.self_check_files = matched_files

        # 6️⃣ Clear listbox (danh sách file .java sẽ được append từ nhiều file self-check)
        self.file_listbox.delete(0, tk.END)

        # 7️⃣ Gọi process_selfcheck_excel cho từng file self-check
        for path in self.self_check_files:
            try:
                process_selfcheck_excel(
                    file_path=path,
                    label_widget=self.self_check_label,
                    listbox_widget=self.file_listbox,
                    screen_code_entry=self.screen_code_entry,
                    author_entry=self.author_entry,
                    clear_listbox=False,  # rất quan trọng: giữ lại các file đã add trước đó
                )
            except Exception as e:
                self.output_text.insert(
                    tk.END, f"\n⚠️ Lỗi khi xử lý {os.path.basename(path)}: {e}\n"
                )

        # 8️⃣ Sau khi đã có đầy đủ danh sách file .java trong listbox, gọi count_code
        try:
            total_code, total_blank, total_comment = count_code(
                listbox_widget=self.file_listbox,
                output_widget=self.output_text,
            )
            # (count_code đã tự xóa nội dung output_widget trước khi ghi)
            self.output_text.insert(
                tk.END,
                f"\n==== SUMMARY FROM FOLDER ====\n"
                f"📂 Thư mục: {folder_path}\n"
                f"🧾 Code: {total_code}, ␣ Blank: {total_blank}, 🗒️ Comment: {total_comment}\n",
            )
        except Exception as e:
            self.output_text.insert(
                tk.END,
                f"\n⚠️ Lỗi khi đếm dòng code từ danh sách file: {e}\n",
            )

    def export_selfcheck_report(self):
        if not getattr(self, "self_check_files", None):
            messagebox.showwarning(
                "Thiếu dữ liệu",
                "Chưa có danh sách file self-check. Hãy chọn thư mục Self Check trước.",
            )
            return

        rows = []
        for sc_path in self.self_check_files:
            base = os.path.basename(sc_path)

            # trích mã màn hình từ tên file
            screen = ""
            parts = base.split("_")
            for p in parts:
                up = p.upper()
                if up.startswith("GUI") and len(up) >= 8 and up[3:8].isdigit():
                    screen = up[:8]  # GUI + 5 số
                    break

            # 🔹 Lấy danh sách source .java 状態=新規 bằng process_selfcheck_excel
            sources = process_selfcheck_excel(
                file_path=sc_path,
                label_widget=self.self_check_label,  # hoặc 1 label khác nếu bạn muốn
                listbox_widget=self.file_listbox,  # có thể là listbox UI thật
                screen_code_entry=self.screen_code_entry,
                author_entry=self.author_entry,
                clear_listbox=False,  # rất quan trọng: không xóa listbox UI
            )

            if not sources:
                # Không có file .java 新規 → vẫn ghi dòng với 0
                rows.append(
                    {
                        "Màn hình": screen,
                        "File self-check": base,
                        "Số file": 0,
                        "Dòng code": 0,
                        "Dòng trắng": 0,
                        "Dòng comment": 0,
                    }
                )
                continue

            # 🔹 Tạo Listbox tạm chỉ để feed count_code
            temp_listbox = tk.Listbox()
            for src in sources:
                temp_listbox.insert(tk.END, src)

            # 🔹 Text tạm – không gán parent để tránh attribute error
            temp_output = tk.Text()

            # Sử dụng count_code
            total_code, total_blank, total_comment = count_code(
                listbox_widget=temp_listbox,
                output_widget=temp_output,
            )

            rows.append(
                {
                    "Màn hình": screen,
                    "File self-check": base,
                    "Số file": len(sources),
                    "Dòng code": total_code,
                    "Dòng trắng": total_blank,
                    "Dòng comment": total_comment,
                }
            )

        if not rows:
            messagebox.showwarning("Không có dữ liệu", "Không tạo được dòng nào để xuất.")
            return

        df = pd.DataFrame(
            rows,
            columns=[
                "Màn hình",
                "File self-check",
                "Số file",
                "Dòng code",
                "Dòng trắng",
                "Dòng comment",
            ],
        )

        save_path = filedialog.asksaveasfilename(
            title="Lưu báo cáo",
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")],
            initialfile="selfcheck_report.xlsx",
        )
        if not save_path:
            return

        df.to_excel(save_path, index=False)
        self._style_xlsx(save_path, df)

        messagebox.showinfo("Hoàn tất", f"Đã xuất báo cáo: {save_path}")

    def _style_xlsx(self, path_xlsx: str, df: pd.DataFrame):
        """Định dạng file .xlsx: header đậm, auto-filter, freeze, căn lề, #,##0, auto-width."""

        wb = load_workbook(path_xlsx)
        ws = wb.active

        # Freeze header + AutoFilter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Style header
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill
            cell.border = border

        # Căn lề & định dạng số cho các cột số
        num_cols = {"Số file", "Dòng code", "Dòng trắng", "Dòng comment"}
        {c: i + 1 for i, c in enumerate(df.columns)}

        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                # viền mảnh nhẹ
                cell.border = border

                # nếu là cột số -> #,##0 và căn phải; còn lại căn trái
                col_header = ws.cell(row=1, column=cell.column).value
                if col_header in num_cols:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto width theo nội dung (giới hạn tối đa để không quá rộng)
        max_width = 80
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_text = str(ws.cell(row=1, column=col_idx).value or "")
            width = len(header_text) + 2

            for row_idx in range(2, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                s = f"{v}"
                # cột số thường không cần quá dài
                if header_text in num_cols:
                    width = max(width, len(s))
                else:
                    width = max(width, len(s))

            ws.column_dimensions[col_letter].width = min(width + 2, max_width)

        wb.save(path_xlsx)
