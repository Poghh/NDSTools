import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from toolsAction.beActions.count_code import count_code
from toolsAction.beActions.process_selfcheck_excel import (
    process_selfcheck_excel,
)
from toolsUI.beTab.unit_test_generater_dialog import UnitTestDialog

from .subTabs.comment_and_unit_test_tab import CommentAndUnitTestTab
from .subTabs.dto_and_db_tab import DtoAndDbTab
from .subTabs.self_check_tab import SelfCheckTab

PRIMARY_COLOR = "#2563eb"  # xanh primary
PRIMARY_DARK = "#1d4ed8"
SIDEBAR_BG = "#111827"
TEXT_MUTED = "#6b7280"
APP_BG = "#f3f4f6"


class BackEndTab:
    def __init__(self, parent):
        self.tab = tk.Frame(parent, bg=APP_BG)
        parent.add(self.tab, text="Back-End")

        # style dùng trong build_selfcheck_frame (lookup background)
        self.style = ttk.Style()

        self.build_ui()

    def build_ui(self):
        # === LAYOUT CHÍNH CỦA TAB: sidebar (col 0) + content (col 1) ===
        self.tab.columnconfigure(0, weight=0)  # sidebar
        self.tab.columnconfigure(1, weight=1)  # nội dung chính
        self.tab.rowconfigure(0, weight=1)

        # ===================== SIDEBAR BÊN TRÁI =====================
        sidebar = tk.Frame(self.tab, bg=SIDEBAR_BG)
        sidebar.grid(row=0, column=0, sticky="nsw")
        sidebar.rowconfigure(99, weight=1)  # đẩy khoảng trống xuống dưới

        tk.Label(
            sidebar,
            text="Menu",
            bg=SIDEBAR_BG,
            fg="#e5e7eb",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", pady=(10, 6), padx=10)

        self.workflow_var = tk.StringVar(value="selfcheck")

        def make_side_btn(text, name):
            return tk.Button(
                sidebar,
                text=text,
                anchor="w",
                bd=0,
                relief="flat",
                bg=SIDEBAR_BG,
                fg="#e5e7eb",
                activebackground=PRIMARY_DARK,
                activeforeground="#ffffff",
                padx=14,
                pady=8,
                font=("Segoe UI", 9),
                highlightthickness=0,
                # -> gọi show_workflow, trong đó tự đổi màu sidebar
                command=lambda: self.show_workflow(name),
            )

        self.btn_sc = make_side_btn("🧾  Self Check", "selfcheck")
        self.btn_sc.pack(fill="x")

        self.btn_cm = make_side_btn("📝  Comment & Unit Test", "comment")
        self.btn_cm.pack(fill="x")

        self.btn_dev = make_side_btn("🛠  DTO & DB Tools", "devtools")
        self.btn_dev.pack(fill="x")

        tk.Label(
            sidebar,
            text="Review Tool v2.0",
            bg=SIDEBAR_BG,
            fg=TEXT_MUTED,
            font=("Segoe UI", 8),
        ).pack(anchor="w", pady=(10, 4), padx=12, side="bottom")

        # ===================== CONTENT BÊN PHẢI =====================
        # giống ui_demo: main -> container_card (workflow) + notebook_card (kết quả)
        main = ttk.Frame(self.tab, style="TFrame")
        main.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        main.rowconfigure(0, weight=0)  # workflow (trên)
        main.rowconfigure(1, weight=1)  # notebook (dưới)
        main.columnconfigure(0, weight=1)

        # --- Card chứa các workflow frame ---
        container_card = ttk.Frame(main, style="Card.TFrame")
        container_card.grid(row=0, column=0, sticky="ew")
        container_card.columnconfigure(0, weight=1)

        self.workflow_container = ttk.Frame(container_card, style="Card.TFrame")
        self.workflow_container.grid(row=0, column=0, sticky="ew")
        self.workflow_container.columnconfigure(0, weight=1)

        # --- Card chứa Notebook kết quả ---
        notebook_card = ttk.Frame(main, style="Card.TFrame")
        notebook_card.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        notebook_card.rowconfigure(0, weight=1)
        notebook_card.columnconfigure(0, weight=1)

        self.output_notebook = ttk.Notebook(notebook_card)
        self.output_notebook.grid(row=0, column=0, sticky="nsew")

        # tạo 4 tab kết quả (Log / Self Check Result / Comment & UT / DTO & DB)
        self.build_output_tabs()

        # === BUILD 3 WORKFLOW FRAMES (giống ui_demo) ===
        self.selfcheck_tab = SelfCheckTab(
            parent=self.workflow_container,
            style=self.style,
            # action=self.fake_action,
            # fake_paste_codes=self.fake_paste_codes,
        )
        self.comment_and_unit_test_tab = CommentAndUnitTestTab(
            parent=self.workflow_container, style=self.style
        )
        self.dto_and_db_tab = DtoAndDbTab(parent=self.workflow_container, style=self.style)

        # Ẩn 2 tab còn lại lúc khởi tạo
        self.comment_and_unit_test_tab.frame.grid_remove()
        self.dto_and_db_tab.frame.grid_remove()

        # Mặc định hiển thị Self Check
        self.show_workflow("selfcheck")

    # ------------------------------------------------------------------ #
    # CHUYỂN GIỮA CÁC WORKFLOW (SELF CHECK / COMMENT / DTO & DB)
    # ------------------------------------------------------------------ #
    def show_workflow(self, name):
        # Ẩn tất cả frame workflow
        for t in (
            self.selfcheck_tab,
            self.comment_and_unit_test_tab,
            self.dto_and_db_tab,
        ):
            t.frame.grid_remove()

        # Hiện đúng frame
        if name == "selfcheck":
            self.selfcheck_tab.frame.grid(row=0, column=0, sticky="nsew")
            # nếu bạn có status_var thì set, còn không thì bỏ dòng này đi
            # self.status_var.set("Workflow: Self Check")
        elif name == "comment":
            self.comment_and_unit_test_tab.frame.grid(row=0, column=0, sticky="nsew")
            # self.status_var.set("Workflow: Comment & Unit Test")
        elif name == "devtools":
            self.dto_and_db_tab.frame.grid(row=0, column=0, sticky="nsew")
            # self.status_var.set("Workflow: DTO & DB Tools")

        # Cập nhật màu sidebar
        self._set_sidebar_active(name)

    def _set_sidebar_active(self, active_name: str):
        """Đổi màu nút sidebar theo workflow đang chọn (giống ui_demo)."""
        normal_bg = SIDEBAR_BG
        normal_fg = "#e5e7eb"
        active_bg = PRIMARY_COLOR
        active_fg = "#ffffff"

        buttons = [
            ("selfcheck", self.btn_sc),
            ("comment", self.btn_cm),
            ("devtools", self.btn_dev),
        ]

        for name, btn in buttons:
            if name == active_name:
                btn.configure(
                    bg=active_bg,
                    fg=active_fg,
                    # font=("Segoe UI", 9, "bold"),
                )
            else:
                btn.configure(
                    bg=normal_bg,
                    fg=normal_fg,
                    # font=("Segoe UI", 9),
                )

    # ------------------------------------------------------------------ #
    # OUTPUT NOTEBOOK
    # ------------------------------------------------------------------ #
    def build_output_tabs(self):
        # Tab Log
        tab_log = ttk.Frame(self.output_notebook)
        self.output_notebook.add(tab_log, text="Log")

        self.log_text = scrolledtext.ScrolledText(
            tab_log, wrap=tk.WORD, borderwidth=0, background="#ffffff"
        )
        self.log_text.pack(fill="both", expand=True)

        # Tab Self Check Result
        tab_sc = ttk.Frame(self.output_notebook)
        self.output_notebook.add(tab_sc, text="Self Check Result")

        self.sc_result_text = scrolledtext.ScrolledText(
            tab_sc, wrap=tk.WORD, borderwidth=0, background="#ffffff"
        )
        self.sc_result_text.pack(fill="both", expand=True)

        # Tab Comment & UT
        tab_cm = ttk.Frame(self.output_notebook)
        self.output_notebook.add(tab_cm, text="Comment & Unit Test")

        self.cm_result_text = scrolledtext.ScrolledText(
            tab_cm, wrap=tk.WORD, borderwidth=0, background="#ffffff"
        )
        self.cm_result_text.pack(fill="both", expand=True)

        # Tab DTO / DB
        tab_dev = ttk.Frame(self.output_notebook)
        self.output_notebook.add(tab_dev, text="DTO / DB")

        self.dev_result_text = scrolledtext.ScrolledText(
            tab_dev, wrap=tk.WORD, borderwidth=0, background="#ffffff"
        )
        self.dev_result_text.pack(fill="both", expand=True)

    def check_and_open_unittest_dialog(self):
        screen_code = self.screen_code_entry.get().strip()

        if not screen_code:
            messagebox.showwarning(
                "Thiếu thông tin", "Vui lòng nhập đầy đủ Tên tác giả và Mã màn hình."
            )
            return

        UnitTestDialog(self.tab, screen_code=screen_code)

    def select_self_check_folder(self):
        """Chọn thư mục chứa các file Self Check, lọc theo mã GUI trong ô văn bản và xử lý từng file Excel."""
        # 1️⃣ Lấy danh sách mã GUIxxxxx từ vùng văn bản
        raw = self.screen_codes_text.get("1.0", tk.END)
        found = re.findall(
            r"(?<![A-Za-z0-9])(GUI\d{5}|[A-Z][A-Z0-9]{5})(?![A-Za-z0-9])", raw, flags=re.IGNORECASE
        )

        seen = set()
        self.screen_codes = []
        for c in (x.upper().strip() for x in found):
            if c and c not in seen:
                seen.add(c)
                self.screen_codes.append(c)

        if not self.screen_codes:
            messagebox.showinfo(
                "Chưa có danh sách mã",
                "Vui lòng dán danh sách mã (mỗi dòng 1 mã) vào ô phía trên trước khi chọn thư mục.",
            )
            return

        # 2️⃣ Chọn thư mục chứa file self-check
        folder_path = filedialog.askdirectory(title="Chọn thư mục Self Check")
        if not folder_path:
            return

        self.self_check_path = folder_path
        self.self_check_label.config(text=f"Thư mục: {os.path.basename(folder_path)}")

        # 3️⃣ Lấy danh sách file Excel trong thư mục
        valid_exts = (".xlsx", ".xls", ".xlsm", ".csv")
        all_files = [
            f
            for f in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, f)) and f.lower().endswith(valid_exts)
        ]
        if not all_files:
            messagebox.showwarning(
                "Không có file",
                "Không tìm thấy file Self Check (.xlsx/.xls/.xlsm/.csv) trong thư mục này.",
            )
            return

        # 4️⃣ Lọc file có chứa mã
        matched_files = []
        for f in all_files:
            fu = f.upper()
            if any(code in fu for code in self.screen_codes):
                matched_files.append(os.path.join(folder_path, f))

        if not matched_files:
            messagebox.showwarning(
                "Không tìm thấy", "Không có file nào trùng với mã màn hình đã dán."
            )
            return

        # 5️⃣ Lưu danh sách vào biến instance
        self.self_check_files = matched_files

        # 6️⃣ Clear listbox (danh sách file .java sẽ được append từ nhiều file self-check)
        self.file_listbox.delete(0, tk.END)

        # 7️⃣ Gọi process_selfcheck_excel cho từng file self-check
        for path in self.self_check_files:
            try:
                process_selfcheck_excel(
                    file_path=path,
                    label_widget=self.self_check_label,
                    listbox_widget=self.file_listbox,
                    screen_code_entry=self.screen_code_entry,
                    author_entry=self.author_entry,
                    clear_listbox=False,  # rất quan trọng: giữ lại các file đã add trước đó
                )
            except Exception as e:
                self.output_text.insert(
                    tk.END, f"\n⚠️ Lỗi khi xử lý {os.path.basename(path)}: {e}\n"
                )

        # 8️⃣ Sau khi đã có đầy đủ danh sách file .java trong listbox, gọi count_code
        try:
            total_code, total_blank, total_comment = count_code(
                listbox_widget=self.file_listbox,
                output_widget=self.output_text,
            )
            # (count_code đã tự xóa nội dung output_widget trước khi ghi)
            self.output_text.insert(
                tk.END,
                f"\n==== SUMMARY FROM FOLDER ====\n"
                f"📂 Thư mục: {folder_path}\n"
                f"🧾 Code: {total_code}, ␣ Blank: {total_blank}, 🗒️ Comment: {total_comment}\n",
            )
        except Exception as e:
            self.output_text.insert(
                tk.END,
                f"\n⚠️ Lỗi khi đếm dòng code từ danh sách file: {e}\n",
            )

    def export_selfcheck_report(self):
        if not getattr(self, "self_check_files", None):
            messagebox.showwarning(
                "Thiếu dữ liệu",
                "Chưa có danh sách file self-check. Hãy chọn thư mục Self Check trước.",
            )
            return

        rows = []
        for sc_path in self.self_check_files:
            base = os.path.basename(sc_path)

            # trích mã màn hình từ tên file
            screen = ""
            parts = base.split("_")
            for p in parts:
                up = p.upper()
                if up.startswith("GUI") and len(up) >= 8 and up[3:8].isdigit():
                    screen = up[:8]  # GUI + 5 số
                    break

            # 🔹 Lấy danh sách source .java 状態=新規 bằng process_selfcheck_excel
            sources = process_selfcheck_excel(
                file_path=sc_path,
                label_widget=self.self_check_label,  # hoặc 1 label khác nếu bạn muốn
                listbox_widget=self.file_listbox,  # có thể là listbox UI thật
                screen_code_entry=self.screen_code_entry,
                author_entry=self.author_entry,
                clear_listbox=False,  # rất quan trọng: không xóa listbox UI
            )

            if not sources:
                # Không có file .java 新規 → vẫn ghi dòng với 0
                rows.append(
                    {
                        "Màn hình": screen,
                        "File self-check": base,
                        "Số file": 0,
                        "Dòng code": 0,
                        "Dòng trắng": 0,
                        "Dòng comment": 0,
                    }
                )
                continue

            # 🔹 Tạo Listbox tạm chỉ để feed count_code
            temp_listbox = tk.Listbox()
            for src in sources:
                temp_listbox.insert(tk.END, src)

            # 🔹 Text tạm – không gán parent để tránh attribute error
            temp_output = tk.Text()

            # Sử dụng count_code
            total_code, total_blank, total_comment = count_code(
                listbox_widget=temp_listbox,
                output_widget=temp_output,
            )

            rows.append(
                {
                    "Màn hình": screen,
                    "File self-check": base,
                    "Số file": len(sources),
                    "Dòng code": total_code,
                    "Dòng trắng": total_blank,
                    "Dòng comment": total_comment,
                }
            )

        if not rows:
            messagebox.showwarning("Không có dữ liệu", "Không tạo được dòng nào để xuất.")
            return

        df = pd.DataFrame(
            rows,
            columns=[
                "Màn hình",
                "File self-check",
                "Số file",
                "Dòng code",
                "Dòng trắng",
                "Dòng comment",
            ],
        )

        save_path = filedialog.asksaveasfilename(
            title="Lưu báo cáo",
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")],
            initialfile="selfcheck_report.xlsx",
        )
        if not save_path:
            return

        df.to_excel(save_path, index=False)
        self._style_xlsx(save_path, df)

        messagebox.showinfo("Hoàn tất", f"Đã xuất báo cáo: {save_path}")

    def _style_xlsx(self, path_xlsx: str, df: pd.DataFrame):
        """Định dạng file .xlsx: header đậm, auto-filter, freeze, căn lề, #,##0, auto-width."""

        wb = load_workbook(path_xlsx)
        ws = wb.active

        # Freeze header + AutoFilter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Style header
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill("solid", fgColor="DDDDDD")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill
            cell.border = border

        # Căn lề & định dạng số cho các cột số
        num_cols = {"Số file", "Dòng code", "Dòng trắng", "Dòng comment"}
        {c: i + 1 for i, c in enumerate(df.columns)}

        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in r:
                # viền mảnh nhẹ
                cell.border = border

                # nếu là cột số -> #,##0 và căn phải; còn lại căn trái
                col_header = ws.cell(row=1, column=cell.column).value
                if col_header in num_cols:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto width theo nội dung (giới hạn tối đa để không quá rộng)
        max_width = 80
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            header_text = str(ws.cell(row=1, column=col_idx).value or "")
            width = len(header_text) + 2

            for row_idx in range(2, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                s = f"{v}"
                # cột số thường không cần quá dài
                if header_text in num_cols:
                    width = max(width, len(s))
                else:
                    width = max(width, len(s))

            ws.column_dimensions[col_letter].width = min(width + 2, max_width)

        wb.save(path_xlsx)

    def _set_sidebar_active(self, active_name: str):
        """Đổi màu nút sidebar theo mục đang chọn (chỉ highlight, chưa ẩn/hiện nội dung)."""
        self.workflow_var.set(active_name)

        normal_bg = SIDEBAR_BG
        normal_fg = "#e5e7eb"
        active_bg = PRIMARY_COLOR
        active_fg = "#ffffff"

        buttons = [
            ("selfcheck", self.btn_sc),
            ("comment", self.btn_cm),
            ("devtools", self.btn_dev),
        ]

        for name, btn in buttons:
            if name == active_name:
                btn.configure(bg=active_bg, fg=active_fg, font=("Segoe UI", 9, "bold"))
            else:
                btn.configure(bg=normal_bg, fg=normal_fg, font=("Segoe UI", 9))
