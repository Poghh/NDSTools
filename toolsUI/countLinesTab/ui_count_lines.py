import os
import threading
import tkinter as tk
from tkinter import filedialog, scrolledtext, ttk

import pandas as pd
from tkinterdnd2 import DND_FILES


class CountLinesTab:
    def __init__(self, tab_parent):
        self.tab = ttk.Frame(tab_parent, style="Custom.TFrame")
        tab_parent.add(self.tab, text="Count Lines")
        
        self.folder_path = ""
        self.running = False
        self.results_data = []  # Lưu kết quả tổng hợp màn hình để export excel
        self.file_details_data = []  # Lưu chi tiết từng file để export excel
        
        self.init_ui()

    def init_ui(self):
        # Style configuration
        style = ttk.Style()
        style.configure("Custom.TFrame", background="#f5f7fa")
        style.configure("Custom.TLabelframe", background="#e3eafc", borderwidth=2, relief="ridge")
        style.configure(
            "Custom.TButton",
            font=("Segoe UI", 10),
            padding=6,
            background="#fff",
            foreground="#222",
            borderwidth=1,
            relief="ridge",
        )
        style.map(
            "Custom.TButton",
            background=[("active", "#e3eafc"), ("disabled", "#f0f0f0")],
            foreground=[("active", "#0d47a1"), ("disabled", "#888")],
            bordercolor=[("active", "#2d5be3"), ("!active", "#b0c4de")],
        )

        self.root = self.tab

        # Main container
        main_frame = tk.Frame(self.root, bg="#f5f7fa")
        main_frame.pack(fill="both", expand=True, padx=16, pady=16)

        # Input section
        input_frame = ttk.LabelFrame(
            main_frame,
            text="Cấu hình đếm dòng",
            padding=12,
            style="Custom.TLabelframe",
        )
        input_frame.pack(fill="x", pady=(0, 16))

        # Folder selection
        folder_frame = tk.Frame(input_frame, bg="#e3eafc")
        folder_frame.pack(fill="x", pady=(0, 12))

        tk.Label(
            folder_frame,
            text="Folder chứa file self-check:",
            bg="#e3eafc",
            font=("Segoe UI", 10, "bold"),
        ).pack(side=tk.LEFT, padx=(0, 10))

        self.folder_btn = tk.Button(
            folder_frame,
            text="Chọn folder",
            command=self.select_folder,
            bg="#fff",
            fg="#222",
            font=("Segoe UI", 10),
            relief=tk.RIDGE,
            bd=1,
            activebackground="#e3eafc",
            activeforeground="#0d47a1",
        )
        self.folder_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.folder_label = tk.Label(
            folder_frame,
            text="Chưa chọn folder",
            fg="gray",
            bg="#e3eafc",
            font=("Segoe UI", 10, "italic"),
        )
        self.folder_label.pack(side=tk.LEFT)

        # Screen names input
        screen_frame = tk.Frame(input_frame, bg="#e3eafc")
        screen_frame.pack(fill="both", expand=True, pady=(0, 12))

        tk.Label(
            screen_frame,
            text="Danh sách màn hình (mỗi dòng một màn hình):",
            bg="#e3eafc",
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", pady=(0, 5))

        self.screen_input = scrolledtext.ScrolledText(
            screen_frame,
            height=10,
            wrap=tk.WORD,
            font=("Consolas", 10),
            bg="#fafdff",
            relief=tk.GROOVE,
            borderwidth=2,
        )
        self.screen_input.pack(fill="both", expand=True)

        # Enable drag and drop for screen input
        try:
            drop_register = getattr(self.screen_input, 'drop_target_register', None)
            dnd_bind = getattr(self.screen_input, 'dnd_bind', None)
            if drop_register and dnd_bind:
                drop_register(DND_FILES)
                dnd_bind("<<Drop>>", self.handle_screen_drop)
        except (AttributeError, Exception):
            pass

        # Option checkbox
        option_frame = tk.Frame(input_frame, bg="#e3eafc")
        option_frame.pack(fill="x", pady=(12, 0))

        self.show_details_var = tk.BooleanVar(value=True)  # Mặc định là chi tiết
        self.show_details_checkbox = tk.Checkbutton(
            option_frame,
            text="Hiển thị chi tiết từng file",
            variable=self.show_details_var,
            bg="#e3eafc",
            font=("Segoe UI", 10),
            activebackground="#e3eafc",
        )
        self.show_details_checkbox.pack(side=tk.LEFT)

        # Buttons section
        button_frame = tk.Frame(main_frame, bg="#f5f7fa")
        button_frame.pack(fill="x", pady=(0, 16))

        self.count_btn = tk.Button(
            button_frame,
            text="Đếm dòng",
            command=self.start_count_lines,
            bg="#fff",
            fg="#222",
            font=("Segoe UI", 10, "bold"),
            relief=tk.RIDGE,
            bd=1,
            activebackground="#e3eafc",
            activeforeground="#0d47a1",
        )
        self.count_btn.pack(side=tk.LEFT, padx=(0, 16))

        self.clear_btn = tk.Button(
            button_frame,
            text="Clear",
            command=self.clear_all,
            bg="#fff",
            fg="#222",
            font=("Segoe UI", 10, "bold"),
            relief=tk.RIDGE,
            bd=1,
            activebackground="#ffebee",
            activeforeground="#d32f2f",
        )
        self.clear_btn.pack(side=tk.LEFT, padx=(0, 16))

        # Export Excel button
        self.export_btn = tk.Button(
            button_frame,
            text="Export Excel",
            command=self.export_to_excel,
            bg="#fff",
            fg="#222",
            font=("Segoe UI", 10, "bold"),
            relief=tk.RIDGE,
            bd=1,
            activebackground="#e3eafc",
            activeforeground="#0d47a1",
        )
        self.export_btn.pack(side=tk.LEFT, padx=(0, 16))

        # Progress bar
        self.progress = ttk.Progressbar(button_frame, mode="indeterminate", length=200)
        self.progress.pack(side=tk.LEFT, padx=(16, 0))
        self.progress.pack_forget()

        # Results section
        results_frame = ttk.LabelFrame(
            main_frame,
            text="Kết quả đếm dòng",
            padding=8,
            style="Custom.TLabelframe",
        )
        results_frame.pack(fill="both", expand=True)

        self.output_text = scrolledtext.ScrolledText(
            results_frame,
            wrap=tk.WORD,
            font=("Consolas", 10),
            bg="#fafdff",
            relief=tk.GROOVE,
            borderwidth=2,
        )
        self.output_text.pack(fill="both", expand=True)

    def select_folder(self):
        folder = filedialog.askdirectory(title="Chọn folder chứa file self-check")
        if folder:
            self.folder_path = folder
            self.folder_label.config(text=folder, fg="green")

    def handle_screen_drop(self, event):
        try:
            file_path = event.data
            if file_path.startswith("{"):
                file_path = file_path[1:-1]

            if file_path.lower().endswith((".txt", ".csv")):
                with open(file_path, encoding="utf-8") as f:
                    content = f.read()
                self.screen_input.insert("1.0", content)
            else:
                # If it's not a text file, just insert the filename
                screen_name = os.path.basename(file_path).replace(".xlsx", "").replace(".xls", "")
                self.screen_input.insert(tk.END, f"{screen_name}\n")
        except Exception as e:
            self.output_text.insert(tk.END, f"Lỗi khi xử lý file: {str(e)}\n")

    def clear_all(self):
        self.screen_input.delete("1.0", tk.END)
        self.output_text.delete("1.0", tk.END)
        self.folder_path = ""
        self.folder_label.config(text="Chưa chọn folder", fg="gray")
        self.results_data = []
        self.file_details_data = []

    def start_count_lines(self):
        if self.running:
            return
        
        screen_names_text = self.screen_input.get("1.0", tk.END).strip()
        if not screen_names_text:
            self.output_text.insert(tk.END, "Vui lòng nhập danh sách màn hình\n")
            return

        if not self.folder_path or not os.path.exists(self.folder_path):
            self.output_text.insert(tk.END, "Vui lòng chọn folder chứa file self-check\n")
            return

        self.output_text.delete("1.0", tk.END)
        self.running = True
        self.count_btn.config(state=tk.DISABLED)
        self.progress.pack(side=tk.LEFT, padx=(16, 0))
        self.progress.start()
        
        threading.Thread(target=self._count_lines_thread, daemon=True).start()

    def _count_lines_thread(self):
        try:
            self.count_lines_from_screens()
        finally:
            self.running = False
            self.count_btn.config(state=tk.NORMAL)
            self.progress.stop()
            self.progress.pack_forget()

    def count_lines_from_screens(self):
        screen_names_text = self.screen_input.get("1.0", tk.END).strip()
        screen_names = [name.strip() for name in screen_names_text.split("\n") if name.strip()]

        # Reset results data
        self.results_data = []
        self.file_details_data = []

        for screen_name in screen_names:
            self.output_text.insert(tk.END, f"\n{'='*60}\n")
            self.output_text.insert(tk.END, f"Màn hình: {screen_name}\n")
            self.output_text.update()

            # Tìm file self-check tương ứng
            excel_file = self.find_selfcheck_file(screen_name)
            if not excel_file:
                self.output_text.insert(tk.END, f"Không tìm thấy file self-check cho màn hình: {screen_name}\n")
                # Lưu kết quả lỗi
                self.results_data.append({
                    "Màn hình": screen_name,
                    "File self-check": "Không tìm thấy",
                    "Số file": 0,
                    "Dòng code": 0,
                    "Dòng trắng": 0,
                    "Dòng comment": 0,
                })
                continue

            excel_filename = os.path.basename(excel_file)
            self.output_text.insert(tk.END, f"File self-check: {excel_filename}\n")

            # Đọc file self-check và lấy các path
            file_paths = self.extract_paths_from_selfcheck(excel_file)
            if not file_paths:
                self.output_text.insert(tk.END, f"Không tìm thấy path nào có trạng thái '新規'\n")
                # Lưu kết quả không có file
                self.results_data.append({
                    "Màn hình": screen_name,
                    "File self-check": excel_filename,
                    "Số file": 0,
                    "Dòng code": 0,
                    "Dòng trắng": 0,
                    "Dòng comment": 0,
                })
                continue

            self.output_text.insert(tk.END, f"Tìm thấy {len(file_paths)} file\n")

            # Kiểm tra xem có hiển thị chi tiết không
            show_details = self.show_details_var.get()

            # Đếm dòng từ các file
            screen_code = 0
            screen_blank = 0
            screen_comment = 0

            # Chi tiết từng file
            for file_path in file_paths:
                code, blank, comment = self.count_file_lines(file_path)
                screen_code += code
                screen_blank += blank
                screen_comment += comment

                # Lưu chi tiết từng file (luôn lưu để export Excel)
                file_name = os.path.basename(file_path) if os.path.exists(file_path) else file_path
                self.file_details_data.append({
                    "Màn hình": screen_name,
                    "File path": file_path,
                    "Tên file": file_name,
                    "Dòng code": code,
                    "Dòng trắng": blank,
                    "Dòng comment": comment,
                })

                # Hiển thị chi tiết từng file (chỉ khi checkbox được chọn)
                if show_details:
                    self.output_text.insert(
                        tk.END,
                        f"  {file_name}: {code} dòng code, {blank} dòng trắng, {comment} dòng comment\n"
                    )
                    self.output_text.update()

            # Thêm dòng trống nếu không hiển thị chi tiết
            if not show_details:
                self.output_text.insert(tk.END, "\n")

            # Lưu kết quả tổng hợp
            self.results_data.append({
                "Màn hình": screen_name,
                "File self-check": excel_filename,
                "Số file": len(file_paths),
                "Dòng code": screen_code,
                "Dòng trắng": screen_blank,
                "Dòng comment": screen_comment,
            })

            self.output_text.insert(
                tk.END,
                f"\n{screen_name}: {screen_code} dòng code, {screen_blank} dòng trắng, {screen_comment} dòng comment\n"
            )
            self.output_text.update()

        self.output_text.insert(tk.END, "\nĐếm dòng hoàn tất.\n")

    def find_selfcheck_file(self, screen_name):
        """Tìm file self-check tương ứng với tên màn hình"""
        if not os.path.exists(self.folder_path):
            return None

        # Tìm file Excel có chứa tên màn hình trong tên file
        for root, dirs, files in os.walk(self.folder_path):
            for file in files:
                if file.lower().endswith((".xlsx", ".xls")):
                    # Kiểm tra nếu tên màn hình có trong tên file
                    if screen_name.lower() in file.lower() or file.lower().startswith(screen_name.lower()):
                        return os.path.join(root, file)

        return None

    def extract_paths_from_selfcheck(self, excel_file):
        """Đọc file self-check và trả về danh sách các path có cột 3 = '新規'"""
        try:
            workbook = pd.read_excel(excel_file, sheet_name="機能別ソース一覧", header=None)
        except Exception as e:
            self.output_text.insert(tk.END, f"Lỗi đọc sheet 機能別ソース一覧: {str(e)}\n")
            return []

        paths = []
        for _index, row in workbook.iterrows():
            try:
                # Cột 3 (index 3) = 状態 (tình trạng), Cột 2 (index 2) = ファイルパス (đường dẫn file)
                if pd.notna(row.iloc[3]) and str(row.iloc[3]).strip() == "新規":
                    if pd.notna(row.iloc[2]):
                        path = str(row.iloc[2]).strip()
                        if path:
                            paths.append(path)
            except (IndexError, Exception):
                continue

        return paths

    def count_file_lines(self, file_path):
        """Đếm số dòng code, blank, comment từ một file"""
        code_count = 0
        comment_count = 0
        blank_count = 0

        # Kiểm tra file có tồn tại không
        if not os.path.exists(file_path):
            return code_count, blank_count, comment_count

        try:
            with open(file_path, encoding="utf-8") as f:
                lines = f.readlines()

            for line in lines:
                trimmed = line.strip()
                if trimmed.startswith(("//", "/*", "*", "<!--")):
                    comment_count += 1
                elif trimmed == "":
                    blank_count += 1
                else:
                    code_count += 1

        except Exception as e:
            # Nếu không đọc được file, bỏ qua
            pass

        return code_count, blank_count, comment_count

    def export_to_excel(self):
        """Export kết quả ra file Excel"""
        if not self.results_data:
            self.output_text.insert(tk.END, "Không có dữ liệu để export. Vui lòng chạy đếm dòng trước.\n")
            return

        # Chọn nơi lưu file
        file_path = filedialog.asksaveasfilename(
            title="Lưu file Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )

        if not file_path:
            return

        try:
            # Tạo DataFrame từ kết quả tổng hợp
            df_summary = pd.DataFrame(self.results_data)

            # Ghi ra Excel
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Sheet tổng hợp
                df_summary.to_excel(writer, sheet_name="Tổng hợp màn hình", index=False)

                # Sheet chi tiết từng file (nếu có)
                if self.file_details_data:
                    df_details = pd.DataFrame(self.file_details_data)
                    df_details.to_excel(writer, sheet_name="Chi tiết từng file", index=False)

                # Format cột cho sheet tổng hợp
                from openpyxl.utils import get_column_letter
                worksheet_summary = writer.sheets["Tổng hợp màn hình"]
                for idx, col in enumerate(df_summary.columns, 1):
                    col_letter = get_column_letter(idx)
                    worksheet_summary.column_dimensions[col_letter].width = 30

                # Format cột cho sheet chi tiết (nếu có)
                if self.file_details_data:
                    df_details = pd.DataFrame(self.file_details_data)
                    worksheet_details = writer.sheets["Chi tiết từng file"]
                    
                    # Format độ rộng cột
                    for idx, col in enumerate(df_details.columns, 1):
                        col_letter = get_column_letter(idx)
                        if col == "File path":
                            worksheet_details.column_dimensions[col_letter].width = 50
                        else:
                            worksheet_details.column_dimensions[col_letter].width = 30
                    
                    # Merge cột "Màn hình" khi có cùng giá trị
                    from openpyxl.styles import Alignment, Border, Side
                    screen_col_idx = df_details.columns.get_loc("Màn hình") + 1  # +1 vì Excel bắt đầu từ 1
                    screen_col_letter = get_column_letter(screen_col_idx)
                    
                    current_screen = None
                    start_row = 2  # Bắt đầu từ dòng 2 (dòng 1 là header)
                    
                    for row_idx, (_, row) in enumerate(df_details.iterrows(), start=0):
                        screen = row["Màn hình"]
                        excel_row = row_idx + 2  # +2 vì Excel bắt đầu từ 1 và có header
                        
                        if current_screen != screen:
                            # Nếu có nhóm trước đó, merge nó
                            if current_screen is not None and start_row < excel_row:
                                worksheet_details.merge_cells(
                                    f"{screen_col_letter}{start_row}:{screen_col_letter}{excel_row - 1}"
                                )
                                # Căn giữa cho cell đã merge
                                cell = worksheet_details[f"{screen_col_letter}{start_row}"]
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                # Thêm border
                                thin_border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                                cell.border = thin_border
                            
                            # Bắt đầu nhóm mới
                            current_screen = screen
                            start_row = excel_row
                    
                    # Merge nhóm cuối cùng
                    if current_screen is not None:
                        end_row = len(df_details) + 1
                        if start_row < end_row:
                            worksheet_details.merge_cells(
                                f"{screen_col_letter}{start_row}:{screen_col_letter}{end_row}"
                            )
                            cell = worksheet_details[f"{screen_col_letter}{start_row}"]
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border

            self.output_text.insert(tk.END, f"\nĐã export kết quả ra file: {file_path}\n")
        except Exception as e:
            self.output_text.insert(tk.END, f"Lỗi khi export Excel: {str(e)}\n")

